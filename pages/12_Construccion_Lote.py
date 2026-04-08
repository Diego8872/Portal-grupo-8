import streamlit as st
import base64
import json
import pandas as pd
import openpyxl
import xlrd
from xlutils.copy import copy as xl_copy
import pdfplumber
import re
import io
import os
from datetime import datetime

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;600&display=swap');
html, body, [class*="css"] { font-family: 'IBM Plex Sans', sans-serif; }
.stApp { background-color: #0f1117; }
h1, h2, h3 { font-family: 'IBM Plex Mono', monospace; color: #f0f0f0; }
.titulo-app { font-family: 'IBM Plex Mono', monospace; font-size: 2rem; font-weight: 600; color: #FFD600; border-bottom: 2px solid #FFD600; padding-bottom: 0.5rem; margin-bottom: 1.5rem; }
.paso-header { font-family: 'IBM Plex Mono', monospace; font-size: 0.85rem; color: #FFD600; letter-spacing: 0.1em; text-transform: uppercase; margin-bottom: 0.3rem; }
.info-box { background: #1a1f2e; border: 1px solid #2a2f3e; border-left: 3px solid #FFD600; padding: 0.8rem 1rem; border-radius: 4px; margin: 0.5rem 0; font-size: 0.9rem; color: #c0c0c0; }
.alerta-box { background: #1f1510; border: 1px solid #5a3a1a; border-left: 3px solid #FF8C00; padding: 0.8rem 1rem; border-radius: 4px; margin: 0.5rem 0; font-size: 0.9rem; color: #ffb060; }
.ok-box { background: #101f15; border: 1px solid #1a5a2a; border-left: 3px solid #00c853; padding: 0.8rem 1rem; border-radius: 4px; margin: 0.5rem 0; font-size: 0.9rem; color: #60d080; }
.stButton > button { background-color: #FFD600; color: #0f1117; font-family: 'IBM Plex Mono', monospace; font-weight: 600; border: none; border-radius: 4px; padding: 0.6rem 1.5rem; }
.stButton > button:hover { background-color: #ffe033; color: #0f1117; }
div[data-baseweb="select"] > div { background-color: #1a1f2e; border-color: #2a2f3e; color: #f0f0f0; }
.stRadio label { color: #c0c0c0 !important; }
.stRadio > div { color: #c0c0c0 !important; }
[data-testid="stToolbar"] { visibility: hidden !important; }
[data-testid="stDecoration"] { display: none !important; }
a[href*="github.com"] { display: none !important; }
div[data-testid="stExpander"] { background-color: #1a1f2e !important; border: 1px solid #2a2f3e !important; border-radius: 6px !important; }
div[data-testid="stExpander"] summary { color: #f0f0f0 !important; }
div[data-testid="stExpander"] > div > div { background-color: #1a1f2e !important; }
label { color: #c0c0c0 !important; }
.stTextInput input { background-color: #1a1f2e !important; color: #f0f0f0 !important; border-color: #2a2f3e !important; }
p, span, div { color: #c0c0c0; }
</style>
""", unsafe_allow_html=True)

PAISES = {
    "BRASIL": 203, "BRAZIL": 203, "ESTADOS UNIDOS": 212, "USA": 212, "UNITED STATES": 212, "US": 212,
    "REINO UNIDO": 426, "UNITED KINGDOM": 426, "UK": 426, "GB": 426, "GREAT BRITAIN": 426,
    "ALEMANIA": 438, "GERMANY": 438, "DEUTSCHLAND": 438, "DE": 438,
    "ITALIA": 417, "ITALY": 417, "IT": 417, "AUSTRIA": 405, "AT": 405,
    "SUECIA": 429, "SWEDEN": 429, "SE": 429, "NORUEGA": 422, "NORWAY": 422, "NO": 422,
    "FINLANDIA": 411, "FINLAND": 411, "FI": 411,
    "REPUBLICA CHECA": 451, "REP. CHECA": 451, "CZECH REPUBLIC": 451, "CZ": 451,
    "TAIWAN": 313, "TW": 313, "CHINA": 310, "CN": 310, "VIETNAM": 337, "VN": 337,
    "COREA DEL SUR": 309, "SOUTH KOREA": 309, "KOREA": 309,
    "JAPON": 320, "JAPAN": 320, "INDIA": 315, "SUIZA": 430, "SWITZERLAND": 430,
    "PAISES BAJOS": 423, "NETHERLANDS": 423, "HOLLAND": 423,
    "FRANCE": 412, "FRANCIA": 412, "FR": 412, "ESPAÑA": 410, "SPAIN": 410, "ES": 410, "BELGICA": 406, "BELGIUM": 406, "BE": 406,
    "CANADA": 204, "MEXICO": 218, "MÉXICO": 218, "ARGENTINA": 200, "CHILE": 208,
    "PERU": 222, "PERÚ": 222, "COLOMBIA": 205, "SINGAPORE": 333, "SINGAPUR": 333,
    "MALAYSIA": 326, "MALASIA": 326, "INDONESIA": 316, "THAILAND": 335, "TAILANDIA": 335,
    "TURQUIA": 436, "TURKEY": 436, "ISRAEL": 319, "SUDAFRICA": 159, "SOUTH AFRICA": 159,
    "AUSTRALIA": 501, "NUEVA ZELANDA": 504, "NEW ZEALAND": 504,
    "POLONIA": 424, "POLAND": 424, "HUNGRIA": 414, "HUNGARY": 414, "RUMANIA": 427, "ROMANIA": 427,
    "BULGARIA": 407, "CROACIA": 447, "CROATIA": 447, "ESLOVAQUIA": 448, "SLOVAKIA": 448,
    "ESLOVENIA": 449, "SLOVENIA": 449, "RUSIA": 444, "RUSSIA": 444, "UCRANIA": 445, "UKRAINE": 445,
    "PORTUGAL": 425, "GRECIA": 413, "GREECE": 413, "LUXEMBURGO": 419, "LUXEMBOURG": 419,
    "IRLANDA": 415, "IRELAND": 415, "DINAMARCA": 409, "DENMARK": 409, "HONG KONG": 341,
}

UNIDAD_CME = {
    "KG": 1, "KILOGRAM": 1, "KILOGRAMO": 1, "MT": 2, "M": 2, "METRO": 2, "METROS": 2,
    "M2": 3, "M²": 3, "M3": 4, "M³": 4, "L": 5, "LT": 5, "LITRO": 5, "LITROS": 5,
    "PC": 7, "UNI": 7, "UNIDAD": 7, "UNIT": 7, "PCS": 7, "UN": 7,
    "PAR": 8, "PAIR": 8, "DOC": 9, "DOCENA": 9,
    "G": 14, "GR": 14, "GRAMO": 14, "GRAMOS": 14, "GRAM": 14,
    "TON": 29, "TONELADA": 29, "ML": 47, "MILILITRO": 47,
}

UNIDAD_SIDOM = {
    1: "01 - KILOGRAMO", 2: "02 - METROS", 3: "03 - METRO CUADRADO", 4: "04 - METRO CUBICO",
    5: "05 - LITROS", 7: "07 - UNIDAD", 8: "08 - PAR", 9: "09 - DOCENA",
    14: "14 - GRAMO", 29: "29 - TONELADA", 47: "47 - MILILITRO",
}

def get_codigo_pais(texto):
    if not texto: return None
    t = str(texto).strip().upper()
    # Match exacto primero (para códigos ISO de 2 letras)
    if t in PAISES: return PAISES[t]
    # Match parcial para nombres completos
    for k, v in PAISES.items():
        if len(k) > 2 and (k in t or t in k): return v
    return None

def get_codigo_unidad(texto):
    if not texto: return 7
    return UNIDAD_CME.get(str(texto).strip().upper(), 7)

def limpiar_numero(texto):
    if texto is None: return 0.0
    try:
        s = str(texto).strip()
        if ',' in s and '.' in s:
            s = s.replace('.','').replace(',','.') if s.rfind(',') > s.rfind('.') else s.replace(',','')
        elif ',' in s: s = s.replace(',','.')
        return float(s)
    except: return 0.0

def extraer_items_natura(texto_pdf):
    """
    FIX: campo VENTA opcional con (?:\\d+\\s+)?
    Soporta:
      50293720 1.050,000 KG descripcion...     (sin VENTA)
      50391678 118997 828 PC descripcion...    (con VENTA)
    """
    items = []
    patron = re.compile(
        r'^(\d{5,8})\s+(?:\d+\s+)?([\d.,]+)\s+(KG|G|PC|MT|L|LT|M)\s+(.+?)\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s*$',
        re.IGNORECASE
    )
    skip = ["NCM","NALADI","VOLUMES","Pallet","TOTAL","PESO NETO","PESO BRUTO",
            "CANTIDAD CAJAS","FLETE","SEGURO","MEDIDA","CONDICIÓN","OBSERVACIÓN",
            "MARCA","IMPORTADOR","RECEPTOR","BENEFICIARIO","DESCRIPCIÓN"]
    for l in texto_pdf.split('\n'):
        l = l.strip()
        if any(s in l for s in skip) or l.startswith("---"): continue
        m = patron.match(l)
        if m:
            cant = limpiar_numero(m.group(2))
            total = limpiar_numero(m.group(7))
            unitario = limpiar_numero(m.group(6))
            if unitario == 0 and cant > 0: unitario = round(total / cant, 2)
            items.append({
                "codigo": m.group(1).strip(), "descripcion": m.group(4).strip(),
                "cantidad": cant, "unidad_cod": get_codigo_unidad(m.group(3).upper()),
                "unidad_raw": m.group(3).upper(), "peso_neto": limpiar_numero(m.group(5)),
                "unitario": unitario, "total": total, "origen": 203, "procedencia": 203, "moneda": "ARS",
            })
    return items

def extraer_items_hci(texto_pdf):
    items = []
    lineas = texto_pdf.split('\n')
    origen = 203
    for i, l in enumerate(lineas):
        if "ORIGEM" in l.upper() or "ORIGIN" in l.upper():
            for j in range(i, min(i+5, len(lineas))):
                cod = get_codigo_pais(lineas[j])
                if cod: origen = cod; break
    patron = re.compile(r'^(\d+)\s+([\d.,]+)\s+([\d.,]+)\s+(.+?)\s+([\w_]+)\s+(\w+)\s+([\d.,]+)\s+([\d.,]+)$')
    for l in lineas:
        m = patron.match(l.strip())
        if m:
            origen_item = get_codigo_pais(m.group(6).strip()) or origen
            items.append({
                "codigo": m.group(5).strip(), "descripcion": m.group(4).strip(),
                "cantidad": limpiar_numero(m.group(2)), "unidad_cod": 7, "unidad_raw": "PC",
                "peso_neto": limpiar_numero(m.group(3)), "unitario": limpiar_numero(m.group(7)),
                "total": limpiar_numero(m.group(8)), "origen": origen_item, "procedencia": origen_item, "moneda": "USD",
            })
    return items

def extraer_items_wartsila(texto_pdf):
    # FIX: ítem partido en dos líneas
    lineas = texto_pdf.split('\n')
    pat_item   = re.compile(r'^(\d{6})\s+(.+)$')
    pat_datos  = re.compile(r'^(\w+)\s+\w+\s+([\d.,]+)\s+PC\s+([\d.,]+)\s+\w+\s+[\d.,]+%\s+([\d.,]+)')
    pat_origen = re.compile(r'^([A-Z]{2})\s+\d+\s+([\d.,]+)\s+KG')
    items = []
    i = 0
    while i < len(lineas):
        m1 = pat_item.match(lineas[i].strip())
        if m1 and i+1 < len(lineas):
            m2 = pat_datos.match(lineas[i+1].strip())
            if m2:
                # línea 1: "000100 12V92F BATTERY" → grupo1=000100, grupo2="12V92F BATTERY"
                # El Part no. es la primera palabra del grupo2, la descripción el resto
                partes = m1.group(2).strip().split(None, 1)
                codigo = partes[0]
                desc   = partes[1] if len(partes) > 1 else partes[0]
                cant     = limpiar_numero(m2.group(2))
                unitario = limpiar_numero(m2.group(3))
                total    = limpiar_numero(m2.group(4))
                origen_iso = None; peso = 0.0
                for j in range(i+2, min(i+5, len(lineas))):
                    mo = pat_origen.match(lineas[j].strip())
                    if mo: origen_iso = mo.group(1); peso = limpiar_numero(mo.group(2)); break
                origen = get_codigo_pais(origen_iso) or 0
                items.append({
                    "codigo": codigo, "descripcion": desc,
                    "cantidad": cant, "unidad_cod": 7, "unidad_raw": "PC",
                    "peso_neto": peso, "unitario": unitario, "total": total,
                    "origen": origen, "procedencia": origen, "moneda": "EUR",
                })
                i += 2; continue
        i += 1
    return items

def extraer_items_aesa_desde_excel(marcas_bytes):
    items = []
    try:
        df = pd.read_excel(io.BytesIO(marcas_bytes), sheet_name="Pos", dtype=str, skiprows=3)
        df = df[df["Pos"].str.match(r"^\d+$", na=False)].copy()
        for _, row in df.iterrows():
            cod = str(row.get("Código SAP del Material","")).strip()
            cant = limpiar_numero(row.get("Cantidad",0))
            if not cod or cant == 0: continue
            origen = get_codigo_pais(str(row.get("Origen","brasil")).strip()) or 203
            items.append({
                "codigo": cod, "descripcion": str(row.get("Descripción","")).strip().replace("\n"," "),
                "cantidad": cant, "unidad_cod": get_codigo_unidad(str(row.get("Unidad","UNI")).strip().upper()),
                "unidad_raw": str(row.get("Unidad","UNI")).strip().upper(),
                "peso_neto": limpiar_numero(row.get("Peso Neto Total de Posición (kg)",0)),
                "unitario": limpiar_numero(row.get("Precio Unitario de Posición",0)),
                "total": limpiar_numero(row.get("Precio Total de Posición",0)),
                "origen": origen, "procedencia": origen, "moneda": "USD",
            })
    except: pass
    return items

def extraer_items_groq_vision(pdf_bytes):
    try:
        from pdf2image import convert_from_bytes
        from groq import Groq
        groq_key = st.secrets.get("GROQ_API_KEY","")
        if not groq_key: return []
        client = Groq(api_key=groq_key)
        images = convert_from_bytes(pdf_bytes, dpi=200)
        todos_items = []
        prompt = """Analizá esta factura comercial y extraé SOLO los ítems de la tabla de productos.
Para cada ítem retorná un objeto JSON con estos campos exactos:
- codigo: el código REFI CLI (ej: O__CAP01331691, F__BL_04294047). Si no hay columna REFI CLI, usá el part number o código de producto. NUNCA uses la descripción como código.
- descripcion: descripción del producto
- cantidad: número (columna QUANT o QTY)
- unidad: unidad de medida (PC, KG, MT, G, L)
- peso_neto: peso neto en KG
- unitario: precio unitario
- total: precio total
- origen: país de origen (ej: BRASIL)
Retorná ÚNICAMENTE un array JSON válido, sin markdown, sin texto adicional."""
        for img in images:
            buf = io.BytesIO()
            img.save(buf, format="JPEG", quality=90)
            img_b64 = base64.b64encode(buf.getvalue()).decode()
            response = client.chat.completions.create(
                model="meta-llama/llama-4-scout-17b-16e-instruct",
                messages=[{"role":"user","content":[
                    {"type":"image_url","image_url":{"url":f"data:image/jpeg;base64,{img_b64}"}},
                    {"type":"text","text":prompt}
                ]}], max_tokens=3000)
            raw = response.choices[0].message.content.strip().replace("```json","").replace("```","").strip()
            for it in json.loads(raw):
                todos_items.append({
                    "codigo": str(it.get("codigo","")).strip(), "descripcion": str(it.get("descripcion","")).strip(),
                    "cantidad": float(it.get("cantidad",0)), "unidad_cod": get_codigo_unidad(str(it.get("unidad","PC"))),
                    "unidad_raw": str(it.get("unidad","PC")).upper(), "peso_neto": float(it.get("peso_neto",0)),
                    "unitario": float(it.get("unitario",0)), "total": float(it.get("total",0)),
                    "origen": get_codigo_pais(str(it.get("origen","BRASIL"))) or 203,
                    "procedencia": get_codigo_pais(str(it.get("origen","BRASIL"))) or 203, "moneda": "USD",
                })
        return todos_items
    except: return []

def extraer_texto_pdf(pdf_bytes):
    texto = ""
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t: texto += t + "\n"
    return texto

def extraer_items_pdf(pdf_bytes, proveedor_detectado=None):
    texto = extraer_texto_pdf(pdf_bytes)
    if len(texto.strip()) < 50:
        return "escaneado_vision", extraer_items_groq_vision(pdf_bytes), ""
    if "REFI CLI" in texto or "HCI" in texto.upper():
        return "hci", extraer_items_hci(texto), texto
    elif "Wärtsilä" in texto or "Wartsila" in texto:
        return "wartsila", extraer_items_wartsila(texto), texto
    elif "MATERIAL" in texto and "VENTA" in texto and "CANTIDAD" in texto and "Natura" in texto:
        return "natura", extraer_items_natura(texto), texto
    else:
        return "generico", extraer_items_hci(texto), texto

def leer_ncm(ncm_file_bytes, nombre_archivo):
    try:
        if nombre_archivo.endswith(('.xlsm','.xlsx')):
            try:
                df = pd.read_excel(io.BytesIO(ncm_file_bytes), sheet_name="Catálogo", dtype=str, skiprows=2)
                return dict(zip(df["Código del artículo"].str.strip(), df["NCM"].str.strip()))
            except: pass
            try:
                df = pd.read_excel(io.BytesIO(ncm_file_bytes), sheet_name="Hoja1", dtype=str)
                cols = df.columns.tolist()
                cod_col = next((c for c in cols if "art" in c.lower() or "cod" in c.lower() or "material" in c.lower()), cols[0])
                ncm_col = next((c for c in cols if "ncm" in c.lower()), cols[1])
                return dict(zip(df[cod_col].str.strip(), df[ncm_col].str.strip()))
            except: pass
        df = pd.read_excel(io.BytesIO(ncm_file_bytes), dtype=str)
        cols = df.columns.tolist()
        # Detectar formato Wärtsilä: columnas PART NUMBER y PA
        if "PART NUMBER" in cols and "PA" in cols:
            result = {}
            for _, row in df.iterrows():
                pn = str(row.get("PART NUMBER","")).strip()
                pa = str(row.get("PA","")).strip()
                if pn and pa and pa != "nan":
                    # Limpiar sufijo WARTSILA del part number
                    pn_clean = re.sub(r'WARTSILA$', '', pn, flags=re.IGNORECASE).strip()
                    result[pn_clean] = pa
                    result[pn] = pa  # también con sufijo por si acaso
            return result
        cod_col = next((c for c in cols if "art" in str(c).lower() or "cod" in str(c).lower()), cols[0])
        ncm_col = next((c for c in cols if "ncm" in str(c).lower()), cols[1])
        return dict(zip(df[cod_col].astype(str).str.strip(), df[ncm_col].astype(str).str.strip()))
    except Exception as e:
        st.warning(f"Error leyendo NCMs: {e}")
        return {}

def leer_marcas_aesa(marcas_file_bytes):
    try:
        df = pd.read_excel(io.BytesIO(marcas_file_bytes), sheet_name="Pos", dtype=str, skiprows=3)
        df = df[df["Pos"].str.match(r'^\d+$', na=False)].copy()
        col_marca = next((c for c in df.columns if "MARCA" in str(c).upper()), None)
        col_cod   = next((c for c in df.columns if "SAP" in str(c).upper() or "Código" in str(c)), None)
        if col_marca and col_cod:
            return dict(zip(df[col_cod].str.strip(), df[col_marca].str.strip()))
    except Exception as e:
        st.warning(f"Error leyendo marcas: {e}")
    return {}

def generar_excel_cme(items, nombre_lote):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Hoja1"
    ws.append(["","Articulo","Descripcion","NCM","Cantidad","Unitario","Total","Origen","Procedencia","Unidad de Venta","Marca","Modelo","PN","Marca/Modelo/Otro"])
    for item in items:
        ws.append(["", item.get("codigo",""), item.get("descripcion",""), item.get("ncm","SIN NCM"),
                   item.get("cantidad",0), item.get("unitario",0), item.get("total",0),
                   item.get("origen",""), item.get("procedencia",""), item.get("unidad_cod",7),
                   item.get("marca","sin marca"), item.get("codigo",""), item.get("peso_neto",0), item.get("marca_modelo_otro","")])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

def generar_excel_sidom(items, nombre_lote):
    posibles = [
        os.path.join(os.path.dirname(__file__), "Lote_SIDOM.xlsx"),
        os.path.join(os.path.dirname(__file__), "..", "Lote_SIDOM.xlsx"),
        "Lote_SIDOM.xlsx",
    ]
    template_path = next((p for p in posibles if os.path.exists(p)), None)
    if not template_path:
        raise FileNotFoundError(f"No se encontró Lote_SIDOM.xlsx. Buscado en: {posibles}")
    wb = openpyxl.load_workbook(template_path); ws = wb.active
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row: cell.value = None
    for i, item in enumerate(items):
        r = i + 2
        ws.cell(r,1).value = i+1; ws.cell(r,2).value = item.get("codigo","")
        ws.cell(r,3).value = item.get("cantidad",0); ws.cell(r,4).value = UNIDAD_SIDOM.get(item.get("unidad_cod",7),"07 - UNIDAD")
        ws.cell(r,5).value = item.get("unitario",0); ws.cell(r,6).value = item.get("total",0)
        ws.cell(r,11).value = item.get("peso_neto",0); ws.cell(r,14).value = item.get("origen_nombre","")
        ws.cell(r,16).value = item.get("estado","2 - NUEVO SIN USO IMPORTADO"); ws.cell(r,17).value = item.get("ncm","")
        ws.cell(r,21).value = "N"; ws.cell(r,22).value = "N"; ws.cell(r,23).value = "N"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

PAISES_INV = {}
for nombre, cod in PAISES.items():
    if cod not in PAISES_INV: PAISES_INV[cod] = nombre

# ─── APP ──────────────────────────────────────────────────────────────────
st.markdown('<div class="titulo-app">📦 CONSTRUCCIÓN DE LOTE</div>', unsafe_allow_html=True)

if "paso" not in st.session_state: st.session_state.paso = 1
if "config" not in st.session_state: st.session_state.config = {}
if "items" not in st.session_state: st.session_state.items = []
if "alertas_marca" not in st.session_state: st.session_state.alertas_marca = []

# PASO 1
with st.expander("⚙️  CONFIGURACIÓN", expanded=(st.session_state.paso == 1)):
    col1, col2 = st.columns(2)
    with col1:
        st.markdown('<div class="paso-header">01 — Número de Referencia</div>', unsafe_allow_html=True)
        nro_ref = st.text_input("Número de referencia del lote", placeholder="Ej: 4508864970", key="nro_ref")
        st.markdown('<div class="paso-header">02 — Sistema</div>', unsafe_allow_html=True)
        sistema = st.radio("Sistema de despacho", ["CME", "SIDOM"], horizontal=True, key="sistema")
        st.markdown('<div class="paso-header">03 — Cliente</div>', unsafe_allow_html=True)
        if sistema == "CME":
            cliente = st.selectbox("Cliente", ["Natura", "AESA", "Otro"], key="cliente_cme")
        else:
            cliente = st.selectbox("Cliente", ["Genérico", "Otro"], key="cliente_sidom")
    with col2:
        tipo_ref = None
        if sistema == "CME" and cliente == "Natura":
            st.markdown('<div class="paso-header">04 — Tipo de Referencia</div>', unsafe_allow_html=True)
            tipo_ref = st.radio("Tipo de referencia", ["ARG", "Otro"], horizontal=True, key="tipo_ref")
        st.markdown('<div class="paso-header">05 — ¿Ítems Usados?</div>', unsafe_allow_html=True)
        tiene_usados = st.radio("¿Esta operación puede tener ítems usados?", ["No", "Sí"], horizontal=True, key="tiene_usados")
    if st.button("CONFIRMAR CONFIGURACIÓN →"):
        if not nro_ref: st.error("Ingresá el número de referencia")
        else:
            st.session_state.config = {"nro_ref":nro_ref,"sistema":sistema,"cliente":cliente,"tipo_ref":tipo_ref,"tiene_usados":tiene_usados=="Sí"}
            st.session_state.paso = 2; st.rerun()

# PASO 2
if st.session_state.paso >= 2:
    cfg = st.session_state.config
    with st.expander("📁  ARCHIVOS", expanded=(st.session_state.paso == 2)):
        st.markdown(f'<div class="info-box">Sistema: <b>{cfg["sistema"]}</b> | Cliente: <b>{cfg["cliente"]}</b> | Referencia: <b>{cfg["nro_ref"]}</b></div>', unsafe_allow_html=True)
        facturas = st.file_uploader("Facturas PDF (una o más)", type=["pdf"], accept_multiple_files=True, key="facturas")
        ncm_file = st.file_uploader("Excel de NCMs", type=["xlsx", "xlsm", "xls"], key="ncm_file")
        marcas_file = None
        if cfg["cliente"] == "AESA":
            marcas_file = st.file_uploader("Excel de Marcas (AESA - solapa Pos)", type=["xlsx", "xlsm", "xls"], key="marcas_file")
        modo_excel = "único"
        if facturas and len(facturas) > 1:
            st.markdown('<div class="paso-header">¿Cómo generar el Excel?</div>', unsafe_allow_html=True)
            modo_excel_sel = st.radio("Modo de generación", ["Un Excel por factura", "Un solo Excel con todas"], key="modo_excel")
            modo_excel = "por_factura" if "por factura" in modo_excel_sel else "único"
        if facturas and ncm_file:
            if cfg["cliente"] == "AESA" and not marcas_file:
                st.warning("⚠️ Falta el Excel de Marcas para AESA")
            elif st.button("PROCESAR FACTURAS →"):
                st.session_state.paso = 3
                st.session_state.config["modo_excel"] = modo_excel
                st.session_state.facturas_data = [(f.name, f.read()) for f in facturas]
                st.session_state.ncm_data = (ncm_file.name, ncm_file.read())
                st.session_state.marcas_data = (marcas_file.name, marcas_file.read()) if marcas_file else None
                st.rerun()

# PASO 3 — FIX delay: placeholder.empty() antes del rerun
if st.session_state.paso >= 3:
    cfg = st.session_state.config
    with st.expander("⚙️  PROCESAMIENTO", expanded=(st.session_state.paso == 3)):
        if st.session_state.paso == 3:
            placeholder = st.empty()
            with placeholder.container():
                ncm_nombre, ncm_bytes = st.session_state.ncm_data
                ncm_dict = leer_ncm(ncm_bytes, ncm_nombre)
                st.markdown(f'<div class="ok-box">✅ NCMs cargados: {len(ncm_dict)} registros</div>', unsafe_allow_html=True)
                marcas_dict = {}
                if st.session_state.marcas_data:
                    m_nombre, m_bytes = st.session_state.marcas_data
                    marcas_dict = leer_marcas_aesa(m_bytes)
                    st.markdown(f'<div class="ok-box">✅ Marcas cargadas: {len(marcas_dict)} registros</div>', unsafe_allow_html=True)
                todos_items = []; facturas_items = {}
                for nombre_fac, pdf_bytes in st.session_state.facturas_data:
                    tipo_fac, items_raw, texto = extraer_items_pdf(pdf_bytes)
                    if cfg["cliente"] == "AESA" and len(items_raw) == 0 and st.session_state.marcas_data:
                        _, m_bytes = st.session_state.marcas_data
                        items_raw = extraer_items_aesa_desde_excel(m_bytes); tipo_fac = "aesa_excel"
                    st.markdown(f'<div class="info-box">📄 {nombre_fac} → {len(items_raw)} ítems detectados (tipo: {tipo_fac})</div>', unsafe_allow_html=True)
                    if items_raw:
                        with st.expander(f"Ver ítems detectados ({len(items_raw)})"):
                            for it in items_raw:
                                st.write(f"`{it.get('codigo')}` | {str(it.get('descripcion',''))[:50]} | cant: {it.get('cantidad')} | total: {it.get('total')}")
                    items_enriquecidos = []; alertas_marca = []; alertas_usados = []
                    for item in items_raw:
                        cod = item["codigo"]
                        item["ncm"] = ncm_dict.get(cod, "SIN NCM")
                        if cfg["cliente"] == "Natura":
                            if cfg["tipo_ref"] == "ARG": item["marca"] = "sin marca"
                            else:
                                desc = item["descripcion"].lower()
                                if "natura" in desc: item["marca"] = "natura"
                                elif "avon" in desc: item["marca"] = "avon"
                                else: item["marca"] = None; alertas_marca.append(item)
                            item["marca_modelo_otro"] = cod
                        elif cfg["cliente"] == "AESA":
                            item["marca"] = marcas_dict.get(cod, "SIN MARCA")
                            if item["marca"] == "SIN MARCA": alertas_marca.append(item)
                            item["marca_modelo_otro"] = ""
                        else:
                            item["marca"] = ""; item["marca_modelo_otro"] = ""
                        item["estado"] = "2 - NUEVO SIN USO IMPORTADO"
                        if cfg["tiene_usados"]:
                            desc_lower = item["descripcion"].lower()
                            if any(p in desc_lower for p in ["used","usad","reman","recondition","rebuilt","gebraucht"]):
                                item["estado"] = "4 - USADO IMPORTADO, INCL. REACOND"; alertas_usados.append(item)
                        item["origen_nombre"] = PAISES_INV.get(item.get("origen"), str(item.get("origen","")))
                        items_enriquecidos.append(item)
                    facturas_items[nombre_fac] = {"items":items_enriquecidos,"alertas_marca":alertas_marca,"alertas_usados":alertas_usados}
                    todos_items.extend(items_enriquecidos)

            # FIX: limpiar antes del rerun para evitar flash
            placeholder.empty()

            st.session_state.todos_items = todos_items
            st.session_state.facturas_items = facturas_items
            st.session_state.alertas_marca_global  = [i for fac in facturas_items.values() for i in fac["alertas_marca"]]
            st.session_state.alertas_usados_global = [i for fac in facturas_items.values() for i in fac["alertas_usados"]]
            st.session_state.paso = 4
            st.rerun()

# PASO 4
if st.session_state.paso >= 4:
    cfg = st.session_state.config
    alertas_marca  = st.session_state.get("alertas_marca_global",[])
    alertas_usados = st.session_state.get("alertas_usados_global",[])
    with st.expander("⚠️  VALIDACIÓN", expanded=True):
        hay_alertas = False
        if alertas_marca:
            hay_alertas = True
            st.markdown(f'<div class="alerta-box">⚠️ {len(alertas_marca)} ítems sin marca detectada — completar antes de generar</div>', unsafe_allow_html=True)
            for item in alertas_marca:
                col1, col2 = st.columns([3,1])
                with col1: st.markdown(f'`{item["codigo"]}` — {item["descripcion"][:80]}')
                with col2:
                    opciones = ["natura","avon","sin marca"] if cfg["cliente"]=="Natura" else ["SIN MARCA"]
                    item["marca"] = st.selectbox("Marca", opciones, key=f"marca_sel_{item['codigo']}_{id(item)}")
                    if cfg["cliente"]=="Natura": item["marca_modelo_otro"] = item["codigo"]
        if alertas_usados:
            hay_alertas = True
            st.markdown(f'<div class="alerta-box">⚠️ {len(alertas_usados)} ítems detectados como posiblemente USADOS</div>', unsafe_allow_html=True)
            for item in alertas_usados:
                col1, col2 = st.columns([3,1])
                with col1: st.markdown(f'`{item["codigo"]}` — {item["descripcion"][:80]}')
                with col2:
                    item["estado"] = st.radio("Estado", ["2 - NUEVO SIN USO IMPORTADO","4 - USADO IMPORTADO, INCL. REACOND"],
                                              key=f"estado_sel_{item['codigo']}_{id(item)}", horizontal=False)
        if cfg["tiene_usados"]:
            st.markdown("---")
            st.markdown("**¿Hay algún ítem adicional que sea USADO y no fue detectado?**")
            todos_codigos = [it["codigo"] for it in st.session_state.todos_items]
            usados_extra = st.multiselect("Seleccionar ítems usados adicionales", options=todos_codigos, key="usados_extra")
            for cod in usados_extra:
                for item in st.session_state.todos_items:
                    if item["codigo"] == cod: item["estado"] = "4 - USADO IMPORTADO, INCL. REACOND"
        if not hay_alertas:
            st.markdown('<div class="ok-box">✅ Sin alertas — todo listo para generar</div>', unsafe_allow_html=True)
        if st.button("GENERAR EXCEL →"):
            st.session_state.paso = 5; st.rerun()

# PASO 5
if st.session_state.paso >= 5:
    cfg = st.session_state.config
    nro_ref = cfg["nro_ref"]; sistema = cfg["sistema"]; modo = cfg.get("modo_excel","único")
    with st.expander("✅  RESULTADO", expanded=True):
        if modo == "por_factura":
            for nombre_fac, fac_data in st.session_state.facturas_items.items():
                items = fac_data["items"]; nombre_archivo = f"Lote_{nro_ref}.xlsx"
                buf = generar_excel_cme(items, nro_ref) if sistema=="CME" else generar_excel_sidom(items, nro_ref)
                st.markdown(f'<div class="ok-box">✅ {nombre_fac} → {len(items)} ítems</div>', unsafe_allow_html=True)
                st.download_button(f"⬇️ Descargar {nombre_archivo}", data=buf, file_name=nombre_archivo,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=f"dl_{nombre_fac}")
        else:
            todos = st.session_state.todos_items; nombre_archivo = f"Lote_{nro_ref}.xlsx"
            buf = generar_excel_cme(todos, nro_ref) if sistema=="CME" else generar_excel_sidom(todos, nro_ref)
            st.markdown(f'<div class="ok-box">✅ {len(todos)} ítems procesados → {nombre_archivo}</div>', unsafe_allow_html=True)
            st.download_button(f"⬇️ Descargar {nombre_archivo}", data=buf, file_name=nombre_archivo,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_unico")
        if st.button("🔄 Nuevo Lote"):
            for key in list(st.session_state.keys()): del st.session_state[key]
            st.rerun()
