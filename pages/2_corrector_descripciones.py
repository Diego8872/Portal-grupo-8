import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import re
import io
import unicodedata
import subprocess
import sys

def instalar(paquete):
    subprocess.check_call([sys.executable, "-m", "pip", "install", paquete, "-q"])

# Groq API para separación y traducción contextual
try:
    from groq import Groq
except ImportError:
    instalar("groq")
    from groq import Groq

def get_ia_model():
    """Inicializa IA (Groq) si hay API key disponible."""
    try:
        api_key = st.secrets["GROQ_API_KEY"]
        if api_key:
            return Groq(api_key=api_key)
    except KeyError:
        pass
    except Exception:
        pass
    return None

# ── Diccionario técnico de repuestos maquinaria pesada ────
DICCIONARIO_TECNICO = {
    # ── Filtros ──
    "filter": "filtro", "filters": "filtros",
    "strainer": "colador", "element": "elemento filtrante",
    "cartridge": "cartucho", "screen": "malla",

    # ── Motor ──
    "engine": "motor", "engines": "motores",
    "piston": "pistón", "pistons": "pistones",
    "cylinder": "cilindro", "cylinders": "cilindros",
    "crankshaft": "cigüeñal", "camshaft": "árbol de levas",
    "conrod": "biela", "liner": "camisa", "liners": "camisas",
    "head": "culata", "heads": "culatas",
    "valve": "válvula", "valves": "válvulas",
    "rocker": "balancín", "rockers": "balancines",
    "lifter": "levantaválvulas", "tappet": "taqué",
    "injector": "inyector", "injectors": "inyectores",
    "nozzle": "tobera", "nozzles": "toberas",
    "turbocharger": "turbo cargador",
    "intake": "admisión", "exhaust": "escape",
    "manifold": "colector", "muffler": "silenciador",
    "starter": "motor de arranque",
    "alternator": "alternador",
    "flywheel": "volante motor",

    # ── Sellos y juntas ──
    "seal": "sello", "seals": "sellos",
    "gasket": "junta", "gaskets": "juntas",
    "o-ring": "junta", "oring": "junta",
    "o´ring": "junta", "o'ring": "junta",
    "retainer": "retén", "retainers": "retenes",
    "wiper": "rascador", "wipers": "rascadores",
    "lip seal": "sello de labio",
    "backup ring": "anillo de respaldo",

    # ── Rodamientos y bujes ──
    "bearing": "rodamiento", "bearings": "rodamientos",
    "bushing": "buje", "bushings": "bujes",
    "sleeve": "manguito", "sleeves": "manguitos",
    "race": "pista", "cone": "cono", "cup": "copa",
    "roller": "rodillo", "rollers": "rodillos",
    "needle": "aguja",

    # ── Transmisión y tren de potencia ──
    "transmission": "transmisión",
    "differential": "diferencial",
    "gearbox": "caja de engranajes",
    "gear": "engranaje", "gears": "engranajes",
    "shaft": "eje", "shafts": "ejes",
    "axle": "eje", "axles": "ejes",
    "coupling": "acople", "couplings": "acoples",
    "clutch": "embrague", "clutches": "embragues",
    "disc": "disco", "discs": "discos",
    "brake": "freno", "brakes": "frenos",
    "sprocket": "rueda dentada", "sprockets": "ruedas dentadas",
    "chain": "cadena", "chains": "cadenas",
    "belt": "correa", "belts": "correas",
    "pulley": "polea", "pulleys": "poleas",
    "idler": "rueda guía",
    "planetary": "planetario",
    "carrier": "portasatélites",
    "sun gear": "engranaje solar",

    # ── Sistema hidráulico ──
    "pump": "bomba", "pumps": "bombas",
    # "motor" no se traduce — genera falsos positivos
    "cylinder": "cilindro", "cylinders": "cilindros",
    "hose": "manguera", "hoses": "mangueras",
    "tube": "tubo", "tubes": "tubos",
    "pipe": "caño", "pipes": "caños",
    "fitting": "fitting", "fittings": "fittings",
    "elbow": "codo", "elbows": "codos",
    "nipple": "niple", "nipples": "niples",
    "flange": "brida", "flanges": "bridas",
    "tee": "tee",
    "adapter": "adaptador", "adapters": "adaptadores",
    "connector": "conector", "connectors": "conectores",
    "relief valve": "válvula de alivio",
    "check valve": "válvula de retención",
    "control valve": "válvula de control",
    "accumulator": "acumulador",
    "reservoir": "depósito",

    # ── Estructura y carrocería ──
    "bracket": "soporte", "brackets": "soportes",
    "plate": "placa", "plates": "placas",
    "cover": "tapa", "covers": "tapas",
    "cap": "tapa", "caps": "tapas",
    "housing": "carcasa", "housings": "carcasas",
    "frame": "bastidor", "frames": "bastidores",
    "guard": "protector", "guards": "protectores",
    "shield": "escudo", "shields": "escudos",
    "panel": "panel", "panels": "paneles",
    "door": "puerta", "doors": "puertas",
    "window": "ventana", "windows": "ventanas",
    "glass": "vidrio",
    "mirror": "espejo", "mirrors": "espejos",

    # ── Fijación y sujeción ──
    "bolt": "perno", "bolts": "pernos",
    "nut": "tuerca", "nuts": "tuercas",
    "washer": "arandela", "washers": "arandelas",
    "screw": "tornillo", "screws": "tornillos",
    "pin": "pasador", "pins": "pasadores",
    "clip": "grapa", "clips": "grapas",
    "ring": "aro", "rings": "aros",
    "snap ring": "anillo elástico",
    "lock": "traba", "locking": "de traba",
    "stud": "espárrago", "studs": "espárragos",

    # ── Suspensión y dirección ──
    "spring": "resorte", "springs": "resortes",
    "shock": "amortiguador", "damper": "amortiguador",
    "tie rod": "barra de dirección",
    "steering": "dirección",
    "knuckle": "mangueta",
    "link": "eslabón", "links": "eslabones",
    "arm": "brazo", "arms": "brazos",
    "rod": "vástago", "rods": "vástagos",

    # ── Tren de rodaje (orugas) ──
    "track": "oruga", "tracks": "orugas",
    "shoe": "zapata", "shoes": "zapatas",
    "pad": "zapata", "pads": "zapatas",
    "link": "eslabón",
    "idler": "rueda guía",
    "roller": "rodillo",
    "blade": "cuchilla", "blades": "cuchillas",
    "bucket": "balde", "buckets": "baldes",
    "boom": "pluma", "booms": "plumas",
    "stick": "brazo", "dipper": "brazo",

    # ── Sistema eléctrico ──
    "switch": "interruptor", "switches": "interruptores",
    "relay": "relé", "relays": "relés",
    "fuse": "fusible", "fuses": "fusibles",
    "sensor": "sensor", "sensors": "sensores",
    "harness": "mazo de cables",
    "wire": "cable", "wires": "cables",
    "lamp": "lámpara", "lamps": "lámparas",
    "gauge": "indicador", "gauges": "indicadores",
    "lever": "palanca", "levers": "palancas",
    "knob": "perilla", "knobs": "perillas",
    "handle": "manija", "handles": "manijas",
    "seat": "asiento", "seats": "asientos",

    # ── Enfriamiento ──
    "radiator": "radiador",
    "fan": "ventilador", "fans": "ventiladores",
    "thermostat": "termostato",
    "cooler": "enfriador", "coolers": "enfriadores",
    "condenser": "condensador",
    "compressor": "compresor",

    # ── Varios ──
    "assembly": "conjunto",
    "kit": "kit",
    "set": "juego",
    "group": "grupo", "gp": "grupo",
    "gp-pr": "grupo de presión",
    "seal_exhaust": "sello de escape",
    "allem": "allen",
    "vibratory": "vibratorio",
    "smooth": "liso",
    "asphalt": "asfalto",
    "single": "simple", "double": "doble",
    "drum": "tambor", "drums": "tambores",
    "pulley": "polea",
    "harness": "mazo de cables",
    "flange": "brida",
}


# ── Lista de palabras españolas para segmentación ─────────
PALABRAS_SEPARACION = set([
    # Preposiciones, artículos, conjunciones
    "de","del","en","con","sin","para","por","al","la","las","los","el","un","una",
    "y","o","a","tipo","uso","entre","sobre","dos","tres","cuatro","sus",
    # Materiales
    "acero","caucho","aluminio","plastico","plasticos","goma","nylon","nitrilo",
    "bronce","cobre","papel","celulosa","tela","malla","laton","fluorocarbono",
    "poliuretano","vulcanizado","vulcanizada","sintetico","textil","metalico","metalica",
    # Componentes
    "filtro","sello","junta","juntas","perno","pernos","tuerca","tuercas","arandela",
    "arandelas","tornillo","tornillos","pasador","pasadores","grapa","aro","aros",
    "anillo","anillos","resorte","resortes","vastago","piston","pistones","cilindro",
    "cilindros","valvula","valvulas","bomba","bombas","motor","motores","eje","ejes",
    "engranaje","engranajes","correa","correas","cadena","cadenas","polea","poleas",
    "rodillo","rodillos","zapata","zapatas","oruga","orugas","eslabon","eslabones",
    "manguera","mangueras","tubo","tubos","brida","bridas","codo","codos","niple",
    "niples","adaptador","adaptadores","conector","conectores","tapa","tapas",
    "cubierta","cubiertas","carcasa","carcasas","soporte","soportes","placa","placas",
    "chapa","chapas","panel","paneles","protector","protectores","brazo","brazos",
    "balde","baldes","cuchilla","cuchillas","palanca","palancas","manija","manijas",
    "perilla","perillas","sensor","sensores","cable","cables","mazo","mazos","radiador",
    "ventilador","ventiladores","buje","bujes","manguito","manguitos","reten","retenes",
    "cojinete","cojinetes","espaciador","espaciadores","laina","lainas","abrazadera",
    "abrazaderas","cardan","bulon","bulones","esparrago","esparragos","clavija",
    "tapon","tapones","interruptor","interruptores","alternador","inyector","inyectores",
    "inyectora","tobera","toberas","acople","acoples","amortiguador","compresor",
    "turbocargador","cartucho","cartuchos","elemento","elementos","colador","maza",
    "conjunto","juego","kit","grupo","oring","ring","segmento","segmentos",
    "suplemento","sujetador","retenedor","pasante","pasantes","orificio","orificios",
    "perforacion","perforaciones","buloneria","planchuela",
    # Adjetivos técnicos
    "hexagonal","rectangular","circular","redondo","ondulado","ondulada","reforzado",
    "reforzada","refuerzos","refuerzo","mecanizado","mecanizada","soldado","soldada",
    "soldadas","plegado","plegada","recto","conico","cilindrico","esferico","delantero",
    "trasero","lateral","superior","inferior","central","interno","interna","externo",
    "hidraulico","hidraulica","hidraulicas","neumatico","electrico","mecanico","termico",
    "simple","doble","triple","radial","axial","macho","hembra","antifriccion","plano",
    "plana","hueco","solido","trenzada","trenzado","giratoria","giratorio","mecanizada",
    "fundicion","tubular","tensionador","guardabarro","guardabarros","barrero","vulcanizada",
    # Técnicos específicos
    "presion","retencion","transmision","freno","embrague","diferencial","alimentacion",
    "conexion","zocalo","zocalos","termoplastico","termoplasticos","extremos","extremo",
    "diametro","longitud","espesor","ajuste","montaje","sujecion","armado","completo",
    "original","izquierdo","derecho","frontal","posterior","primario","secundario",
    "principal","auxiliar","desgaste","proteccion","aleacion","borde","corte","nucleo",
    "intercambiador","calor","bloque","manifold","modulo","tablero","operador","operadores",
    "atenuador","base","tanque","fluido","carrete","cuerpo","plancha","banda","espiral",
    "capot","abulonar","minero","mineros","equipos","cabina","camion","tractor","cargador",
    "excavadora","combustible","aceite","agua","aire","grasa","lubricante","refrigerante",
    "escape","admision","general","industrial","agricola","terminal","terminales","parte",
    "sistema","serie","modelo","endurecer","mando","final","fijo","acondicionado",
    "monta","cabeza","respaldo","espaciadora","metal","comprimido","arranque","bateria",
    "forma","rueda","dentada","mensula","suspension","plataforma","acceso","fondo","luz",
    "controles","control","junto","junto","bornes","borne","batería","junto","controle",
    "junto","pasantes","interno","fluido","plastico",
    # Abreviaciones
    "vcc","vdc","psi","thk","cat","sem",
])



def tiene_palabras_pegadas(texto):
    """Detecta si un texto tiene palabras pegadas."""
    palabras = texto.split()
    # Descripción de una sola palabra larga
    if len(palabras) == 1 and len(texto) > 6:
        return True
    # Alguna palabra larga sin números ni símbolos (probable pegado)
    for p in palabras:
        solo_letras = re.sub(r'[^a-záéíóúüñA-ZÁÉÍÓÚÜÑ]', '', p)
        if len(solo_letras) > 12:
            return True
    # Proporción baja de espacios respecto a la longitud (muchas palabras pegadas)
    if len(texto) > 20 and len(palabras) < len(texto) / 8:
        return True
    return False

def procesar_lote_ia(modelo, descripciones):
    """Manda un lote de descripciones a Gemini para separar y traducir."""
    lista = "\n".join([f"{i+1}. {d}" for i, d in enumerate(descripciones)])
    prompt = f"""Eres un experto en repuestos de maquinaria pesada minera (Caterpillar, Komatsu, Volvo, etc).

Se te dan descripciones de repuestos donde las palabras están pegadas sin espacios. Tu única tarea es insertar los espacios correctos entre las palabras y corregir ortografía básica.

REGLAS:
- Insertar espacios donde corresponde entre palabras
- Traducir palabras en inglés al español: filter→filtro, seal→sello, bearing→rodamiento, housing→carcasa, bracket→soporte, bushing→buje, shaft→eje, bolt→perno, hose→manguera, valve→válvula
- Corregir acentos y ortografía: presion→presión, transmision→transmisión, hidraulico→hidráulico, neumatico→neumático, lubricacion→lubricación
- NO cambiar: marcas (CAT, SEM), modelos (320C), medidas (25MM, 4PSI, VCC), la letra L cuando indica forma

EJEMPLOS (seguir exactamente este estilo):
"Biseladodeplastico" → "Biselado de plástico"
"VALVULADEPASAJEDEAIRECOMPRIMIDO" → "Válvula de pasaje de aire comprimido"
"Filtroelementodepapelcelulosa" → "Filtro elemento de papel celulosa"
"PlacaplanchueladeaceroplegadaenL" → "Placa planchuela de acero plegada en L"
"Hojadecortederecho" → "Hoja de corte derecho"
"BOMBADEPRELUBRICACION" → "Bomba de prelubricación"
"SEGMENTOINTERMEDIODEDIENTEDEPALA" → "Segmento intermedio de diente de pala"

FORMATO DE RESPUESTA — solo el número y el texto corregido, una línea por item:
1. texto corregido aquí
2. texto corregido aquí

Descripciones:
{lista}"""

    try:
        response = modelo.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.1,
        )
        texto_respuesta = response.choices[0].message.content.strip()
        
        lineas = texto_respuesta.split("\n")
        resultados = {}
        for linea in lineas:
            linea = linea.strip()
            if not linea: continue
            # Intentar varios formatos de respuesta
            match = re.match(r'^(\d+)[\.\)\-\:]\s*(.+)$', linea)
            if match:
                idx = int(match.group(1)) - 1
                resultados[idx] = match.group(2).strip()
        return resultados
    except Exception as e:
        return {}

def detectar_equipo_groq(modelo, descripciones):
    """Usa Groq para detectar y separar referencia a equipos en descripciones."""
    lista = "\n".join([f"{i+1}. {d}" for i, d in enumerate(descripciones)])
    prompt = f"""Sos un experto en repuestos de maquinaria pesada (CAT, Komatsu, SEM, Volvo, etc).

Para cada descripción de repuesto, identificá si menciona en qué equipo o máquina se usa.
Devolvé SOLO la referencia al equipo. Si no hay equipo mencionado, devolvé vacío.

REGLAS IMPORTANTES:
- Solo extraer marca/modelo/tipo de equipo (CAT 793, Cargador 972K, Excavadora 320, etc)
- Si el nombre del equipo es PARTE del nombre del repuesto, NO extraer (ej: "Brazo de escobilla de motoniveladora" → vacío)
- Si solo dice "Caterpillar" o "CAT" como marca genérica al final → extraer
- Si dice "equipos varios CAT" o "equipos mineros" → extraer
- Si dice "para freno", "para manguera", "para transmisión" → vacío (son partes, no equipos)
- IMPORTANTE: Responder exactamente el número de items recibidos, en orden

EJEMPLOS:
"Bulón de acero (M10X1.25) Caterpillar" → Caterpillar
"Válvula de solenoide. Topadores D6" → Topadores D6  
"Conjunto del alternador, equipos varios CAT" → equipos varios CAT
"Retén de caucho, mando final de Cargador SEM 636D" → Cargador SEM 636D
"Filtro de aceite de cargador frontal 972K" → cargador frontal 972K
"CINTURON DE SEGURIDAD PARA EQUIPOS MINEROS" → equipos mineros
"acople de INOX, cat 777 CAMION" → cat 777 CAMION
"Sensor de presión Cat de 5V, 48-120 kPa" → 
"BOMBA DE ENGRANAJES PARA FRENO" → 
"BRAZO DE ESCOBILLA DE CABINA DE MOTONIVELADORA" → 
"Turbocargador de motor diesel 3516 de camion 793C" → camion 793C
"Carcasa de acero, parte de transmisión de cargador SEM655" → cargador SEM655
"Medidor de nivel hidráulico. Aplica a Sistema de Equipo Cargador 966" → Equipo Cargador 966
"Correa de caucho para poleas de motor de equipos mineros" → equipos mineros
"ESPACIADOR DE ACERO APLICADO A SISTEMA ELECTRICO DE CAT 990H" → CAT 990H

FORMATO ESTRICTO — exactamente una línea por item, en el mismo orden:
1|equipo o vacío
2|equipo o vacío
...

Descripciones ({len(descripciones)} items):
{lista}"""

    try:
        response = modelo.chat.completions.create(
            model="llama-3.1-8b-instant",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.0,
        )
        texto = response.choices[0].message.content.strip()
        resultados = {}
        VACIOS = {"ninguno","(ninguno)","none","-","vacío","(vacío)","vacio",
                  "(vacio)","","sin equipo","(sin equipo)","n/a","na","no"}
        for linea in texto.split("\n"):
            linea = linea.strip()
            if not linea: continue
            partes = linea.split("|")
            if len(partes) >= 1:
                try:
                    m = re.match(r'(\d+)[|.)]\s*(.*)', linea)
                    if not m: continue
                    idx = int(m.group(1)) - 1
                    equipo = m.group(2).strip() if len(partes) >= 2 else ""
                    equipo = partes[1].strip() if len(partes) >= 2 else ""
                    equipo = "" if equipo.lower().strip("().") in VACIOS else equipo
                    # Validar que el índice sea válido
                    if 0 <= idx < len(descripciones):
                        resultados[idx] = ("", equipo)
                except:
                    pass
        return resultados
    except:
        return {}


def separar_palabras_pegadas(texto, modelo=None):
    """Separa palabras pegadas usando Gemini si disponible, sino diccionario."""
    if not tiene_palabras_pegadas(texto):
        return texto, False
    
    if modelo:
        # Modo Gemini - se llama por lotes desde procesar_archivo
        return texto, False  # placeholder, se procesa en lote
    
    # Fallback: diccionario local
    palabras = texto.split()
    resultado = []
    modificado = False
    for palabra in palabras:
        if (len(palabra) < 8 or
            re.match(r'^[\d\W]+$', palabra) or
            re.match(r'^[A-Z]{1,4}\d+', palabra)):
            resultado.append(palabra)
            continue
        segmentada = segmentar_dp(palabra.lower())
        if segmentada and len(segmentada) > 1:
            if palabra[0].isupper():
                segmentada[0] = segmentada[0].capitalize()
            resultado.append(" ".join(segmentada))
            modificado = True
        else:
            resultado.append(palabra)
    return " ".join(resultado), modificado

def segmentar_dp(texto):
    """Segmentación por programación dinámica - fallback sin Gemini."""
    n = len(texto)
    dp = [None] * (n + 1)
    dp[0] = []
    for i in range(n):
        if dp[i] is None: continue
        for j in range(i + 1, n + 1):
            sub = texto[i:j]
            if sub in PALABRAS_SEPARACION and dp[j] is None:
                dp[j] = dp[i] + [sub]
    return dp[n] if dp[n] is not None else [texto]

PALABRAS_CLAVE = ["KIT", "CONJUNTO", "MANGUERA"]
PALABRAS_CLAVE_FRASE = ["ELEMENTO FILTRANTE"]

CORRECCIONES_ORTOGRAFIA = {
    r'\bpasdor\b': 'pasador',
    r'\bbraxo\b': 'brazo',
    r'\bhidaulico\b': 'hidráulico',
    r'\bhidaulica\b': 'hidráulica',
    r'\bectronico[s]?\b': 'electrónico',
    r'\bectrónico[s]?\b': 'electrónico',
    r'\bdelanero\b': 'delantero',
    r'\bdelanera\b': 'delantera',
    r'\bplastico\b': 'plástico',
    r'\bplastica\b': 'plástica',
    r'\bfundicion\b': 'fundición',
    r'\btransmision\b': 'transmisión',
    r'\bdireccion\b': 'dirección',
    r'\bsuspension\b': 'suspensión',
    r'\binyeccion\b': 'inyección',
    r'\bconexion\b': 'conexión',
    r'\bproteccion\b': 'protección',
    r'\bsujecion\b': 'sujeción',
    r'\brotacion\b': 'rotación',
    r'\bcarcaza\b': 'carcasa',
    r'\bhidraulico\b': 'hidráulico',
    r'\bhidraulica\b': 'hidráulica',
    r'\belectrico\b': 'eléctrico',
    r'\belectrica\b': 'eléctrica',
    r'\bmecanico\b': 'mecánico',
    r'\bneumatico\b': 'neumático',
    r'\bconico\b': 'cónico',
    r'\bconica\b': 'cónica',
    r'\bsintetico\b': 'sintético',
    r'\bsintetica\b': 'sintética',
    r'\bpresion\b': 'presión',
    r'\bvalvula\b': 'válvula',
    r'\bvalvulas\b': 'válvulas',
    r'\bpistons\b': 'pistones',
    r'\bfijacion\b': 'fijación',
    r'\bseparacion\b': 'separación',
    r'\bdistribucion\b': 'distribución',
    r'\bfluoroelastomero\b': 'fluoroelastómero',
}

# ── Funciones ──────────────────────────────────────────────

def limpiar_url(texto):
    # Limpiar caracteres raros de Excel
    texto = texto.replace('_x000D_', '').replace('\xa0', ' ')
    # Limpiar URLs con guión antes
    texto = re.sub(r'\s*-\s*https?://\S+', '', texto)
    texto = re.sub(r'https?://\S+', '', texto)
    return re.sub(r'\s+', ' ', texto).strip()


def normalizar_texto(s):
    """Elimina acentos para comparación."""
    return unicodedata.normalize('NFD', s).encode('ascii', 'ignore').decode('ascii').lower()

EQUIPOS_PALABRAS_SET = {"camion","cargador","excavadora","motoniveladora","topador",
                        "tractor","minicargador","compactador","generador","retroexcavadora",
                        "maquina","equipo","equipos","motor","manipulador",
                        "martillo","motopala","perforadora","volquete"}

EQUIPOS_RE_STR = r'(?:manipulador(?:\s+telescopico)?|camion(?:es)?(?:\s+(?:fuera\s+de\s+ruta|minero))?|camión(?:es)?(?:\s+(?:fuera\s+de\s+ruta|minero))?|cargador(?:a)?(?:\s+frontal)?|excavadora|motoniveladora|topador|tractor|minicargador|compactador|generador|motogenerador|perforadora|retroexcavadora|motopala|martillo|maquina(?:s)?|máquina(?:s)?|equipo(?:s)?|motor(?:\s+(?:diesel|a\s+gas|\d{3,4}[A-Z]?))|cat(?:erpillar)?|sem\b)'
PATRON_PREP_EQUIPO_FINAL = re.compile(
    rf'(?i)(?:,\s*|\s+)[.\s]*\b(?:en|de|para|del)\s+[.]?{EQUIPOS_RE_STR}[\w\s\-\/,.]*$'
)

def limpiar_equipo_de_desc(corregida, equipo_groq):
    """Elimina la referencia al equipo de la descripción, incluyendo la preposición."""
    if not equipo_groq or not corregida: return corregida
    
    # Limpiar trailing preposiciones/puntos sueltos
    desc = re.sub(r'(?i)\s*[.]\s*$', '', corregida.strip())
    desc = re.sub(r'(?i)\s+\b(de|en|para|del|con)\s*[.]?\s*$', '', desc.strip()).strip().rstrip(',.')
    
    # Paso 1: preposición + equipo conocido al final
    match = PATRON_PREP_EQUIPO_FINAL.search(desc)
    if match and match.start() > 5:
        resultado = desc[:match.start()].strip().rstrip(',.').strip()
        if len(resultado) > 5: return resultado
    
    # Paso 2: modelo numérico al final con equipo conocido antes
    m_modelo = re.search(r'(?i)\s+([A-Za-zÁÉÍÓÚáéíóú]+)\s+(?:[A-Z]{1,4})?\d{2,4}[A-Z]?\s*$', desc)
    if m_modelo:
        palabra_antes = normalizar_texto(m_modelo.group(1))
        if palabra_antes in EQUIPOS_PALABRAS_SET:
            idx_corte = m_modelo.start()
            texto_antes = desc[:idx_corte].rstrip()
            mp = re.search(r'(?i)\s+\b(en|de|para|del)\s*$', texto_antes)
            resultado = desc[:mp.start()].strip().rstrip(',.').strip() if mp else texto_antes.strip().rstrip(',.').strip()
            if len(resultado) > 5: return resultado
    
    # Paso 3: modelo solo al final — cortar solo el modelo
    m_eq_modelo = re.search(r'(?:[A-Z]{1,4})?\d{2,4}[A-Z]?$', equipo_groq.strip())
    if m_eq_modelo:
        modelo = m_eq_modelo.group(0)
        if re.search(r'\d', modelo):
            idx = desc.rfind(modelo)
            if idx > 10 and desc[idx:].strip() == modelo:
                resultado = desc[:idx].strip().rstrip(',.').strip()
                if len(resultado) > 5: return resultado
    
    # Paso 4: búsqueda normalizada
    desc_norm = normalizar_texto(desc)
    buscar_norm = normalizar_texto(equipo_groq[:25])
    idx = desc_norm.rfind(buscar_norm)
    if idx > 10:
        texto_antes = desc[:idx].rstrip()
        m = re.search(r'(?i)\s+\b(en|de|para|del|con)\s*[.]?\s*$', texto_antes)
        resultado = desc[:m.start()].strip().rstrip(',.').strip() if m else desc[:idx].strip().rstrip(',.').strip()
        if len(resultado) > 5: return resultado
    
    return desc

def limpiar_codigo_interno(texto):
    """Elimina códigos internos tipo SEAL_EXHAUST_1974834.
    Si el texto es SOLO un código, extrae las palabras descriptivas y las traduce."""
    
    # Detectar si todo el texto es un código tipo PALABRA_PALABRA_NUMERO
    patron_codigo_completo = r'^([A-Z][A-Z_]+)_\d{5,}$'
    match = re.match(patron_codigo_completo, texto.strip())
    
    if match:
        # Extraer partes descriptivas (sin el número final)
        partes = texto.strip().split('_')
        palabras = [p for p in partes if not p.isdigit() and len(p) > 1]
        # Traducir cada parte usando el diccionario
        traducidas = []
        for p in palabras:
            p_lower = p.lower()
            if p_lower in DICCIONARIO_TECNICO:
                traducidas.append(DICCIONARIO_TECNICO[p_lower])
            else:
                traducidas.append(p.capitalize())
        return " ".join(traducidas), True
    
    # Caso normal: eliminar códigos embebidos en el texto
    texto = re.sub(r'\b[A-Z]+_[A-Z]+_\d{5,}\b', '', texto)
    texto = re.sub(r'\b[A-Z_]{3,}_\d{5,}\b', '', texto)
    return re.sub(r'\s+', ' ', texto).strip(), False

def detectar_palabras_clave(texto):
    texto_upper = texto.upper()
    encontradas = []
    # Palabras simples
    for p in PALABRAS_CLAVE:
        if re.search(r'\b' + p + r'\b', texto_upper):
            encontradas.append(p)
    # Frases de dos o más palabras (deben aparecer juntas)
    for frase in PALABRAS_CLAVE_FRASE:
        if frase in texto_upper:
            encontradas.append(frase)
    return " | ".join([f"⚠️ {p}" for p in encontradas]) if encontradas else ""


PATRON_CORTE_EQUIPO = re.compile(
    r'(?i),?[\s.]*(?:'
    r'de\s+motor(?:es)?\s+(?:diesel|cat(?:erpillar)?|a\s+gas|el[eé]ctrico)\b|'
    r'de\s+(?:camion(?:es)?|camión(?:es)?|volquete|cargador(?:a)?|excavadora|motoniveladora|topador|tractor|minicargador|compactador|pavimentador|generador|motogenerador|retroexcavadora|perforadora|maquina|máquina|equipos?)\s+(?:cat\s+)?(?:minero\s+)?\d{2,4}[A-Z]?\b|'
    r'uso\s+en\s+(?:\w+\s+de\s+)?(?:camion(?:es)?|camión(?:es)?|volquete|cargador(?:a)?|excavadora|motoniveladora|topador|tractor|minicargador|compactador|pavimentador|generador|motogenerador|retroexcavadora|perforadora|maquina|máquina|equipos?)|'
    r'uso\s+en\s+[A-Z0-9]{2,6}[A-Z0-9]\b|'
    r'uso\s+en\s+(?:carrocer[ií]a|chasis|grupo|sistema|tren)\s+de\s+(?:camion(?:es)?|camión(?:es)?|volquete|cargador(?:a)?|excavadora|motoniveladora|topador|tractor|minicargador|compactador|pavimentador|generador|motogenerador|retroexcavadora|perforadora|maquina|máquina|equipos?)|'
    r'para\s+equipos?\s+(?:mineros?|cat(?:erpillar)?|varios|barios|industriales?)|'
    r'para\s+motor(?:es)?\s+(?:cat(?:erpillar)?|diesel|a\s+gas)\b|'
    r'para\s+instalaciones?\s+en\s+(?:camion(?:es)?|camión(?:es)?|volquete|cargador(?:a)?|excavadora|motoniveladora|topador|tractor|minicargador|compactador|pavimentador|generador|motogenerador|retroexcavadora|perforadora|maquina|máquina|equipos?)|'
    r'para\s+uso\s+en\s+(?:camion(?:es)?|camión(?:es)?|volquete|cargador(?:a)?|excavadora|motoniveladora|topador|tractor|minicargador|compactador|pavimentador|generador|motogenerador|retroexcavadora|perforadora|maquina|máquina|equipos?|maquinas?\s+cat(?:erpillar)?|mineria)|'
    r'para\s+(?:camion(?:es)?|camión(?:es)?|volquete|cargador(?:a)?|excavadora|motoniveladora|topador|tractor|minicargador|compactador|pavimentador|generador|motogenerador|retroexcavadora|perforadora|maquina|máquina|equipos?)\s*(?:mineros?|cat(?:erpillar)?|marca|fuera\s+de\s+ruta|[0-9]{3,4}[A-Z]?)?(?=\s*[,.\-]|$)|'
    r'para\s+\w+\s+(?:de\s+\w+\s+)?de\s+(?:camion(?:es)?|camión(?:es)?|volquete|cargador(?:a)?|excavadora|motoniveladora|topador|tractor|minicargador|compactador|pavimentador|generador|motogenerador|retroexcavadora|perforadora|maquina|máquina|equipos?)\s+(?:mineros?|cat|caterpillar)|'
    r'uso\s+en\s+(?:maquinas?\s+cat(?:erpillar)?|mineria)|'
    r'uso\s+equipos?\s+(?:varios|barios|mineros?|cat(?:erpillar)?)|'
    r'utilizad[oa]s?\s+en\s+(?:camion(?:es)?|camión(?:es)?|volquete|cargador(?:a)?|excavadora|motoniveladora|topador|tractor|minicargador|compactador|pavimentador|generador|motogenerador|retroexcavadora|perforadora|maquina|máquina|equipos?|cargador\s+marca)|'
    r'instalad[oa]\s+en\s+\w|'
    r'usad[oa]\s+en\s+\w|'
    r'aplicacion\s+(?:camion(?:es)?|camión(?:es)?|volquete|cargador(?:a)?|excavadora|motoniveladora|topador|tractor|minicargador|compactador|pavimentador|generador|motogenerador|retroexcavadora|perforadora|maquina|máquina|equipos?)|'
    r'de\s+(?:cargador(?:a)?|excavadora|motoniveladora|topador)\s+de\s+(?:oruga|ruedas?|orugas?)\s+\d{2,4}[A-Z]?\b|'
    r'de\s+uso\s+en[.\s]+(?:maquinas?\s+cat(?:erpillar)?|mineria|camion(?:es)?|camión(?:es)?|volquete|cargador(?:a)?|excavadora|motoniveladora|topador|tractor|minicargador|compactador|pavimentador|generador|motogenerador|retroexcavadora|perforadora|maquina|máquina|equipos?)|'
    r'(?:camion(?:es)?|camión(?:es)?|volquete|cargador(?:a)?|excavadora|motoniveladora|topador|tractor|minicargador|compactador|pavimentador|generador|motogenerador|retroexcavadora|perforadora|maquina|máquina|equipos?)\s+(?:cat\s+)?(?:minero\s+)?\d{3,4}[A-Z]?\b|'
    r'en\s+equipos?\s+(?:mineros?|cat(?:erpillar)?|varios)|'
    r',?\s+cat\s+\d{3,4}[A-Z]?\s+(?:camion(?:es)?|camión(?:es)?|volquete|cargador(?:a)?|excavadora|motoniveladora|topador|tractor|maquina|máquina)'
    r')')

# Nombres de equipos conocidos para detección directa
def extraer_equipo(texto):
    """Extrae referencia al equipo usando regex."""
    texto = texto.replace('\n', ' ').replace('\r', ' ')
    texto = re.sub(r'\s+', ' ', texto).strip()
    
    match = PATRON_CORTE_EQUIPO.search(texto)
    if match:
        desc_limpia = texto[:match.start()].strip().rstrip(',').strip()
        equipo = texto[match.start():].strip().lstrip(',').strip()
        # Validar que la descripción no quedó vacía o muy corta
        if len(desc_limpia) > 5:
            return desc_limpia, equipo, False
    
    # Marcar para Groq todas las descripciones largas
    necesita_groq = len(texto) > 30
    return texto, "", necesita_groq

def corregir_ortografia(texto):
    errores = []
    for patron, correccion in CORRECCIONES_ORTOGRAFIA.items():
        match = re.search(patron, texto, re.IGNORECASE)
        if match:
            errores.append(f"ortografía: {match.group()}→{correccion}")
            texto = re.sub(patron, correccion, texto, flags=re.IGNORECASE)
    return texto, errores

def es_marca_o_modelo(palabra):
    return bool(re.match(r'^(CAT|Caterpillar|CATERPILLAR|SEM|CAT\d+|[A-Z]{1,4}\d+[A-Z]?|[A-Z]\d+[A-Z]\d*)$', palabra))

def es_medida(palabra):
    return bool(re.match(r'^\d+[\.\-,]?\d*\s*(mm|MM|cm|m|psi|PSI|kg|KG|lb|VCC|VCA|rpm|RPM|pulg|\'|\")?$', palabra))

def traducir_token(token):
    """Traduce solo si la palabra está en el diccionario técnico."""
    limpio = token.strip('.,;:()/\'"`°-').lower()
    if limpio in DICCIONARIO_TECNICO:
        return DICCIONARIO_TECNICO[limpio], limpio
    return None, None


def procesar_descripcion(descripcion_original):
    errores_encontrados = []

    # 1. Limpiar URL
    desc = limpiar_url(descripcion_original)
    if desc != descripcion_original:
        errores_encontrados.append("URL eliminada")

    # 2. Limpiar códigos internos (extraer equipo se hace AL FINAL)
    ref_equipo = ""
    necesita_groq_equipo = False

    # 3. Limpiar códigos internos
    desc_sin_codigos, fue_solo_codigo = limpiar_codigo_interno(desc)
    if fue_solo_codigo:
        errores_encontrados.append(f"código interno traducido: {desc.strip()}→{desc_sin_codigos}")
    elif desc_sin_codigos != desc:
        errores_encontrados.append("código interno eliminado")
    desc = desc_sin_codigos

    # 4. Separar palabras pegadas
    desc_separada, fue_separada = separar_palabras_pegadas(desc)
    if fue_separada:
        errores_encontrados.append("palabras separadas")
    desc = desc_separada

    # 5. Corregir ortografía
    desc, errores_orto = corregir_ortografia(desc)
    errores_encontrados.extend(errores_orto)

    # 6. Traducir palabra por palabra
    tokens = desc.split()
    tokens_nuevos = []
    for token in tokens:
        if es_marca_o_modelo(token) or es_medida(token):
            tokens_nuevos.append(token)
            continue
        traduccion, original = traducir_token(token)
        if traduccion:
            errores_encontrados.append(f"traducido: {original}→{traduccion}")
            tokens_nuevos.append(traduccion)
        else:
            tokens_nuevos.append(token)

    desc = " ".join(tokens_nuevos)

    # 7. Limpiar espacios y normalizar
    desc = re.sub(r'\s+', ' ', desc).strip()
    if desc:
        desc = desc[0].upper() + desc[1:]

    # 8. Detectar palabras clave
    keywords = detectar_palabras_clave(desc)

    # 9. Extraer referencia a equipo (sobre descripción ya corregida)
    desc, ref_equipo, necesita_groq_equipo = extraer_equipo(desc)

    resumen = " | ".join(errores_encontrados) if errores_encontrados else "Sin errores"
    return desc, resumen, keywords, ref_equipo, necesita_groq_equipo


def generar_excel(resultados):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Correcciones"

    header_fill = PatternFill(start_color="1F3A1F", end_color="1F3A1F", fill_type="solid")
    header_font = Font(bold=True, color="B8F542", size=11)
    ok_fill = PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid")
    error_fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
    kw_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    thin = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC')
    )

    encabezados = ["Código", "Descripción Original", "Errores Detectados", "Palabras Clave ⚠️", "Descripción Corregida", "Equipo/Uso"]
    for col, titulo in enumerate(encabezados, 1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    ws.row_dimensions[1].height = 32

    for fila, r in enumerate(resultados, 2):
        for col, val in enumerate([r["codigo"], r["original"], r["errores"], r["keywords"], r["corregida"], r.get("equipo","")], 1):
            cell = ws.cell(row=fila, column=col, value=val)
            cell.border = thin
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        fill = kw_fill if r["keywords"] else (ok_fill if r["errores"] == "Sin errores" else error_fill)
        for col in range(1, 7):
            ws.cell(row=fila, column=col).fill = fill

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 38
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 45
    ws.column_dimensions['F'].width = 55

    ws2 = wb.create_sheet("Resumen")
    total = len(resultados)
    sin_errores = sum(1 for r in resultados if r["errores"] == "Sin errores")
    corregidas = sum(1 for r in resultados if r["errores"] not in ["Sin errores", "Sin descripción"])
    con_kw = sum(1 for r in resultados if r["keywords"])
    ws2['A1'] = "RESUMEN DE PROCESAMIENTO"
    ws2['A1'].font = Font(bold=True, size=14)
    for i, (label, val) in enumerate([("Total:", total), ("Sin errores:", sin_errores), ("Corregidas:", corregidas), ("Palabras clave ⚠️:", con_kw)], 3):
        ws2[f'A{i}'] = label
        ws2[f'B{i}'] = val
    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 10

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ── UI ─────────────────────────────────────────────────────
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Barlow+Condensed:wght@400;600;700;800&family=Barlow:wght@300;400;500&display=swap');

    .stApp { background: linear-gradient(135deg, #0a1628 0%, #0d2137 50%, #0a1628 100%) !important; }
    .block-container { padding-top: 1.5rem !important; max-width: 1100px !important; }

    .hero-wrap {
        background: linear-gradient(90deg, #0d3b6e 0%, #1a5fa8 60%, #0d3b6e 100%);
        border: 1px solid #1e6ab8;
        border-radius: 12px;
        padding: 2.2rem 2.5rem;
        margin-bottom: 2rem;
        position: relative;
        overflow: hidden;
    }
    .hero-wrap::after {
        content: '⚙';
        position: absolute;
        right: 1.5rem; top: 50%;
        transform: translateY(-50%);
        font-size: 9rem;
        opacity: 0.05;
        line-height: 1;
    }
    .hero-tag {
        background: rgba(255,165,0,0.15);
        border: 1px solid rgba(255,165,0,0.4);
        color: #ffb347;
        font-family: 'Barlow Condensed', sans-serif;
        font-size: 0.75rem;
        letter-spacing: 0.2em;
        text-transform: uppercase;
        padding: 3px 14px;
        border-radius: 3px;
        display: inline-block;
        margin-bottom: 0.7rem;
    }
    .hero-title {
        font-family: 'Barlow Condensed', sans-serif;
        font-size: 3rem;
        font-weight: 800;
        color: #fff;
        line-height: 1.05;
        margin-bottom: 0.4rem;
        letter-spacing: 0.01em;
    }
    .hero-title span { color: #ffb347; }
    .hero-sub {
        font-family: 'Barlow', sans-serif;
        font-weight: 300;
        color: rgba(255,255,255,0.55);
        font-size: 0.92rem;
        letter-spacing: 0.04em;
    }
    .hero-badges { margin-top: 1rem; display: flex; gap: 8px; flex-wrap: wrap; }
    .hbadge {
        background: rgba(255,255,255,0.07);
        border: 1px solid rgba(255,255,255,0.12);
        color: rgba(255,255,255,0.7);
        font-family: 'Barlow', sans-serif;
        font-size: 0.78rem;
        padding: 3px 12px;
        border-radius: 20px;
    }

    h1,h2,h3 { font-family: 'Barlow Condensed', sans-serif !important; color: #fff !important; letter-spacing: 0.02em !important; }
    p, .stMarkdown p { color: rgba(255,255,255,0.8) !important; }

    [data-testid="stFileUploader"] {
        background: rgba(13,59,110,0.25) !important;
        border: 2px dashed #1e6ab8 !important;
        border-radius: 10px !important;
    }

    .stButton > button[kind="primary"] {
        background: linear-gradient(90deg, #e67e00, #ffb347) !important;
        color: #0a1628 !important;
        font-family: 'Barlow Condensed', sans-serif !important;
        font-size: 1.15rem !important;
        font-weight: 800 !important;
        letter-spacing: 0.12em !important;
        text-transform: uppercase !important;
        border: none !important;
        border-radius: 8px !important;
        transition: all 0.2s !important;
    }
    .stButton > button[kind="primary"]:hover {
        box-shadow: 0 6px 20px rgba(255,165,0,0.35) !important;
        transform: translateY(-1px) !important;
    }
    .stDownloadButton > button {
        background: linear-gradient(90deg, #1a5fa8, #2980d4) !important;
        color: #fff !important;
        font-family: 'Barlow Condensed', sans-serif !important;
        font-size: 1.1rem !important;
        font-weight: 700 !important;
        letter-spacing: 0.08em !important;
        text-transform: uppercase !important;
        border: none !important;
        border-radius: 8px !important;
    }

    [data-testid="stMetric"] {
        background: rgba(13,59,110,0.4) !important;
        border: 1px solid #1e6ab8 !important;
        border-radius: 10px !important;
        padding: 1rem !important;
    }
    [data-testid="stMetricValue"] { font-family: 'Barlow Condensed', sans-serif !important; color: #ffb347 !important; font-size: 2.2rem !important; }
    [data-testid="stMetricLabel"] { color: rgba(255,255,255,0.55) !important; }

    .stSuccess { background: rgba(0,100,0,0.2) !important; border-left-color: #2d8a2d !important; }
    .stInfo { background: rgba(13,59,110,0.35) !important; border-left-color: #1e6ab8 !important; }
    hr { border-color: rgba(30,106,184,0.25) !important; }
    .stDataFrame { border-radius: 8px !important; }
[data-testid="stToolbar"] { visibility: hidden !important; }
    [data-testid="stDecoration"] { display: none !important; }
    a[href*="github.com"] { display: none !important; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="hero-wrap">
    <div class="hero-tag">🔧 Finning · Repuestos Maquinaria Pesada</div>
    <div class="hero-title">Corrector de<br><span>Descripciones</span></div>
    <div class="hero-sub">Procesamiento inteligente de descripciones de artículos en español</div>
    <div class="hero-badges">
        <span class="hbadge">✓ Traducción semántica</span>
        <span class="hbadge">✓ Corrección ortográfica</span>
        <span class="hbadge">✓ Normalización</span>
        <span class="hbadge">✓ 100% gratuito</span>
    </div>
</div>
""", unsafe_allow_html=True)

st.divider()

archivo = st.file_uploader("📁 Subí tu Excel (Columna A = Código | Columna B = Descripción)", type=["xlsx", "xls"])

if archivo:
    df = pd.read_excel(archivo, header=0)
    df.columns = [str(c).strip() for c in df.columns]

    c1, c2 = st.columns(2)
    with c1:
        st.success(f"✅ **{archivo.name}** — {len(df)} artículos")
    with c2:
        st.info(f"Columnas: **{' | '.join(df.columns.tolist())}**")

    st.dataframe(df.head(5), use_container_width=True)
    st.divider()

    if st.button("▶ Procesar Descripciones", type="primary", use_container_width=True):
        resultados = []
        total = len(df)
        col_codigo = df.columns[0]
        col_desc = df.columns[1]

        # Inicializar IA
        modelo_ia = get_ia_model()
        if modelo_ia:
            st.info("🤖 IA activada — separación y traducción inteligente")
        else:
            st.warning("⚠️ Sin IA — usando modo diccionario")

        progress_bar = st.progress(0)
        status_text = st.empty()
        log_area = st.empty()
        log_lines = []

        # PRE-PROCESO: si hay Gemini, separar en lotes las descripciones pegadas
        descripciones_ia = {}
        if modelo_ia:
            status_text.markdown("🤖 **Paso 1/2:** Procesando con IA...")
            
            # Identificar las que necesitan Gemini
            indices_pegadas = []
            descs_pegadas = []
            for i, row in df.iterrows():
                desc = str(row[col_desc]).strip() if pd.notna(row[col_desc]) else ""
                if desc and desc != "nan" and tiene_palabras_pegadas(desc):
                    indices_pegadas.append(i)
                    descs_pegadas.append(desc)
            
            # Procesar en lotes de 20
            LOTE = 50
            for batch_start in range(0, len(descs_pegadas), LOTE):
                batch_idx = indices_pegadas[batch_start:batch_start+LOTE]
                batch_desc = descs_pegadas[batch_start:batch_start+LOTE]
                
                resultados_gemini = procesar_lote_ia(modelo_ia, batch_desc)
                for j, idx_orig in enumerate(batch_idx):
                    if j in resultados_gemini:
                        descripciones_ia[idx_orig] = resultados_gemini[j]
                
                prog = min((batch_start + LOTE) / max(len(descs_pegadas), 1), 1.0)
                progress_bar.progress(prog * 0.5)
            
            status_text.markdown(f"🤖 IA procesó **{len(descripciones_ia)}** descripciones")
            
            # PASO 1B: detectar equipos con Groq en TODAS las descripciones
            status_text.markdown("🔍 **Paso 1B/2:** Detectando referencias a equipos...")
            indices_equipo = []
            descs_equipo = []
            for i, row in df.iterrows():
                desc = str(row[col_desc]).strip() if pd.notna(row[col_desc]) else ""
                if not desc or desc == "nan": continue
                # Usar descripción ya procesada por IA si existe
                desc_proc = descripciones_ia.get(i, desc)
                indices_equipo.append(i)
                descs_equipo.append(desc_proc)
            
            equipos_groq = {}
            for batch_start in range(0, len(descs_equipo), LOTE):
                batch_idx = indices_equipo[batch_start:batch_start+LOTE]
                batch_desc = descs_equipo[batch_start:batch_start+LOTE]
                resultados_eq = detectar_equipo_groq(modelo_ia, batch_desc)
                for j, idx_orig in enumerate(batch_idx):
                    if j in resultados_eq:
                        equipos_groq[idx_orig] = resultados_eq[j]
                prog_eq = min((batch_start + LOTE) / max(len(descs_equipo), 1), 1.0)
                progress_bar.progress(0.5 + prog_eq * 0.25)
            status_text.markdown(f"🔍 Equipos detectados en **{len(equipos_groq)}** descripciones")

        if not modelo_ia:
            equipos_groq = {}
        
        # PROCESO PRINCIPAL
        status_text.markdown("⚙️ **Paso 2/2:** Aplicando correcciones finales...")
        for i, row in df.iterrows():
            codigo = str(row[col_codigo]).strip()
            desc_original = str(row[col_desc]).strip() if pd.notna(row[col_desc]) else ""

            progress_bar.progress(0.75 + (i + 1) / total * 0.25 if modelo_ia else (i + 1) / total)
            status_text.markdown(f"⚙️ Procesando **{i+1} de {total}**: `{codigo}`")

            if not desc_original or desc_original == "nan":
                resultados.append({"codigo": codigo, "original": "", "errores": "Sin descripción", "keywords": "", "corregida": "", "equipo": ""})
                log_lines.append(f"⬜ [{i+1:03d}] {codigo} → Sin descripción")
            else:
                # Si IA ya procesó esta descripción, usarla como base
                # Usar descripción procesada por IA si existe
                desc_base = descripciones_ia.get(i, desc_original)
                corregida, errores, keywords, equipo, _ = procesar_descripcion(desc_base)
                if i in descripciones_ia and "separado" not in errores.lower():
                    errores = ("separado/traducido por IA | " + errores).rstrip(" | ").replace("Sin errores", "").strip(" | ") or "separado/traducido por IA"
                
                # Groq detecta el equipo — actualizar Equipo/Uso Y limpiar descripción
                if i in equipos_groq:
                    _, equipo_groq_val = equipos_groq[i]
                    if equipo_groq_val and len(equipo_groq_val) < len(corregida):
                        equipo = equipo_groq_val
                        corregida = limpiar_equipo_de_desc(corregida, equipo_groq_val)
                
                # Siempre limpiar URL de la descripción corregida final
                corregida = limpiar_url(corregida)
                # Asegurar capitalización
                if corregida:
                    corregida = corregida[0].upper() + corregida[1:]
                
                resultados.append({"codigo": codigo, "original": desc_original, "errores": errores, "keywords": keywords, "corregida": corregida, "equipo": equipo})
                icono = "⚠️" if keywords else ("✅" if errores == "Sin errores" else "✏️")
                log_lines.append(f"{icono} [{i+1:03d}] {codigo} → {corregida[:55]}...")

            log_area.code("\n".join(log_lines[-12:]), language=None)

        progress_bar.progress(1.0)
        status_text.markdown("✅ **¡Procesamiento completado!**")
        
        st.divider()

        sin_errores = sum(1 for r in resultados if r["errores"] == "Sin errores")
        corregidas = sum(1 for r in resultados if r["errores"] not in ["Sin errores", "Sin descripción"])
        con_kw = sum(1 for r in resultados if r["keywords"])

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Total procesados", total)
        c2.metric("✅ Sin errores", sin_errores)
        c3.metric("✏️ Corregidas", corregidas)
        c4.metric("⚠️ Palabras clave", con_kw)

        st.subheader("📋 Resultados")
        df_out = pd.DataFrame(resultados)
        df_out.columns = ["Código", "Descripción Original", "Errores Detectados", "Palabras Clave", "Descripción Corregida", "Equipo/Uso"]
        st.dataframe(df_out, use_container_width=True, height=420)

        st.divider()
        excel_buffer = generar_excel(resultados)
        st.download_button(
            label="⬇️ Descargar Excel Corregido",
            data=excel_buffer,
            file_name="descripciones_corregidas.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )
