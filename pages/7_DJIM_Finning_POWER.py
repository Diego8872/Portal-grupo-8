import streamlit as st
import pdfplumber
import openpyxl
import subprocess
import os
import re
import datetime
from io import BytesIO
from bs4 import BeautifulSoup


st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;500;600&family=IBM+Plex+Mono:wght@400;500&display=swap');
html, body, [class*="css"] { font-family: 'IBM Plex Sans', sans-serif; }
.djim-header {
    background: linear-gradient(135deg, #1e2440 0%, #2a3060 100%);
    border-radius: 12px; padding: 1.8rem 2.5rem; margin-bottom: 1.5rem;
    border-left: 5px solid #4f8ef7;
}
.djim-header h1 { color: #fff; font-size: 1.7rem; font-weight: 600; margin: 0 0 0.3rem 0; }
.djim-header p { color: #7b8db0; font-size: 0.85rem; margin: 0; font-family: 'IBM Plex Mono', monospace; }
.section-title {
    font-size: 0.7rem; font-weight: 600; letter-spacing: 2px;
    text-transform: uppercase; color: #4f8ef7;
    margin: 2rem 0 0.8rem 0; padding-bottom: 0.5rem; border-bottom: 1px solid #e0e8f0;
}
.alerta-ok {
    background: #e8f5e9; border: 1px solid #a5d6a7; border-radius: 8px;
    padding: 0.8rem 1.2rem; color: #2e7d32; font-weight: 500; font-size: 0.9rem; margin: 0.5rem 0;
}
[data-testid="stToolbar"] { visibility: hidden !important; }
[data-testid="stDecoration"] { display: none !important; }
[data-testid="stHeader"] { display: none !important; }
#GithubIcon { visibility: hidden; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="djim-header">
    <h1>📄 DJIM Finning POWER</h1>
    <p>Generador automático · Interlog Grupo 8</p>
</div>
""", unsafe_allow_html=True)

TEMPLATE_PATH = "template_djim.xlsx"

if "n_items" not in st.session_state:
    st.session_state.n_items = 0

# ─── UTILIDADES PDF ───

def extract_text_pdfplumber(pdf_bytes):
    text = ""
    try:
        with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    text += t + "\n"
    except:
        pass
    return text.strip()


def ocr_pdf_bytes(pdf_bytes, label, dpi=250, psm=None, upscale=1):
    """
    psm: si se especifica, fuerza el modo de segmentación de página de tesseract
         (6 = asume un único bloque uniforme de texto; mejor para tablas chicas
         tipo capturas de pantalla, donde el psm automático reordena columnas).
    upscale: factor de reescalado de la imagen antes de pasarla a tesseract.
             Útil cuando la imagen fuente es de baja resolución (ej: DNRPA
             que son capturas de pantalla), ya que tesseract reconoce mejor
             dígitos y letras chicas en imágenes más grandes.
    """
    tmp_pdf = f"/tmp/{label}.pdf"
    with open(tmp_pdf, "wb") as f:
        f.write(pdf_bytes)
    subprocess.run(["pdftoppm", "-r", str(dpi), tmp_pdf, f"/tmp/ocr_{label}"], capture_output=True)
    images = sorted([x for x in os.listdir("/tmp") if x.startswith(f"ocr_{label}")])
    text = ""
    for img in images:
        img_path = f"/tmp/{img}"
        if upscale and upscale > 1:
            try:
                from PIL import Image, ImageOps
                im = Image.open(img_path).convert("L")
                im = im.resize((im.width * upscale, im.height * upscale), Image.LANCZOS)
                im = ImageOps.autocontrast(im)
                im.save(img_path)
            except Exception:
                pass
        cmd = ["tesseract", img_path, "stdout"]
        if psm:
            cmd += ["--psm", str(psm)]
        result = subprocess.run(cmd, capture_output=True, text=True)
        text += result.stdout
    for img in images:
        try: os.remove(f"/tmp/{img}")
        except: pass
    return text


def get_text(pdf_bytes, label, dpi=250, psm=None, upscale=1):
    text = extract_text_pdfplumber(pdf_bytes)
    chars_utiles = len(re.findall(r'[A-Za-z0-9]', text))
    if not text or chars_utiles < 30:
        text = ocr_pdf_bytes(pdf_bytes, label, dpi=dpi, psm=psm, upscale=upscale)
    return text

def get_text_di(pdf_bytes, label, dpi=250):
    """Para DI: pdfplumber primero, OCR si no tiene CUIT ni fecha."""
    text = extract_text_pdfplumber(pdf_bytes)
    tiene_cuit = bool(re.search(r'\d{2}-\d{8}-\d', text))
    tiene_fecha = bool(re.search(r'\d{2}/\d{2}/\d{4}', text))
    if not tiene_cuit or not tiene_fecha:
        text = ocr_pdf_bytes(pdf_bytes, label, dpi=dpi)
    return text


def normalizar_ocr(text):
    """Corrige errores comunes de OCR antes del parseo."""
    # 1C04 o 1CO4 → IC04 (I confundida con 1, O confundida con 0)
    text = re.sub(r'(?<!\d)1([CG])[O0](\d)', r'IC0\2', text, flags=re.IGNORECASE)
    text = re.sub(r'(?<!\d)1([CG])(\d)', r'I\1\2', text, flags=re.IGNORECASE)
    # ICO4 → IC04
    text = re.sub(r'IC[Oo](\d)', r'IC0\1', text)
    text = re.sub(r'IG[Oo](\d)', r'IG0\1', text)
    # O73 → 073 (O al inicio de número de aduana)
    text = re.sub(r'\bO(\d{2})\b', r'0\1', text)
    return text


# ─── PARSEO DI ───

def parsear_nro_despacho(text_upper):
    m = re.search(r'(\d{2})\s+(\d{3})\s+((?:IC|IG)\d{2})\s+(\d+)\s+([A-Z])(?:\s|$|[^A-Z0-9])', text_upper)
    if m:
        return m.group(1), m.group(2), m.group(3), m.group(4), m.group(5)
    m = re.search(r'(\d{2})\s+(\d{3})\s+([CG]\d{2})\s+(\d+)\s+([A-Z])(?:\s|$|[^A-Z0-9])', text_upper)
    if m:
        return m.group(1), m.group(2), 'I' + m.group(3), m.group(4), m.group(5)
    m = re.search(r'\*(\d{2})(\d{3})(IC\d{2}|IG\d{2})(\d+)([A-Z])\*', text_upper)
    if m:
        return m.groups()
    return None


def parsear_di(text):
    from paises import PAISES
    datos = {}
    alertas = []

    # Normalizar errores de OCR ANTES de procesar
    text_norm = normalizar_ocr(text)
    text_norm_upper = text_norm.upper()

    # Este diccionario ahora se usa SOLO como fallback (ver más abajo), cuando
    # no se pudo parsear el número de despacho. El código numérico de aduana
    # que viene directamente del número de despacho es siempre más confiable
    # que buscar el nombre de la aduana en todo el texto del DI (el nombre
    # puede aparecer también en otras secciones, como la tabla de Ingresos
    # Brutos por jurisdicción, y generar falsos positivos).
    ADUANAS = {
        'BS.AS. (CAPITAL)': '001', 'BS.AS.(CAPITAL)': '001', 'BUENOS AIRES CAPITAL': '001',
        'BAHIA BLANCA': '003', 'BARILOCHE': '004', 'CAMPANA': '008',
        'BARRANQUERAS': '010', 'CLORINDA': '012', 'COLON': '013',
        'COMODORO RIVADAVIA': '014', 'CONCEPCION DEL URUGUAY': '015',
        'CONCORDIA': '016', 'CORDOBA': '017', 'CORRIENTES': '018',
        'PUERTO DESEADO': '019', 'DIAMANTE': '020', 'ESQUEL': '023',
        'FORMOSA': '024', 'GOYA': '025', 'GUALEGUAYCHU': '026',
        'IGUAZU': '029', 'JUJUY': '031', 'LA PLATA': '033',
        'LA QUIACA': '034', 'MAR DEL PLATA': '037', 'MENDOZA': '038',
        'NECOCHEA': '040', 'PARANA': '041', 'PASO DE LOS LIBRES': '042',
        'POCITOS': '045', 'POSADAS': '046', 'PUERTO MADRYN': '047',
        'RIO GALLEGOS': '048', 'RIO GRANDE': '049', 'ROSARIO': '052',
        'SALTA': '053', 'SAN JAVIER': '054', 'SAN JUAN': '055',
        'SAN LORENZO': '057', 'SAN NICOLAS': '059', 'SAN PEDRO': '060',
        'SANTA CRUZ': '061', 'SANTA FE': '062', 'TINOGASTA': '066',
        'USHUAIA': '067', 'VILLA CONSTITUCION': '069', 'EZEIZA': '073',
        'TUCUMAN': '074', 'NEUQUEN': '075', 'ORAN': '076',
        'SAN RAFAEL': '078', 'LA RIOJA': '079', 'SAN ANTONIO OESTE': '080',
        'SAN LUIS': '083', 'SANTO TOME': '084', 'VILLA REGINA': '085',
        'OBERA': '086', 'CALETA OLIVIA': '087', 'GENERAL DEHEZA': '088',
        'SANTIAGO DEL ESTERO': '089', 'GENERAL PICO': '090',
        'BS.AS. NORTE': '091', 'BS.AS. SUR': '092', 'RAFAELA': '093',
        'MULTIADUANA': '099',
    }

    result = parsear_nro_despacho(text_norm_upper)
    if result:
        anio, aduana, tipo, nro, dc = result
        datos['nro_despacho'] = f"{tipo}{nro}{dc}"
        datos['anio'] = anio
        # FIX: el código de aduana (grupo 2) ya viene directamente del propio
        # número de despacho (ej: "26 073 IC04 080265 C" -> 073 = EZEIZA), es
        # confiable y NO se debe pisar con una búsqueda de nombre en todo el
        # texto del DI.
        datos['id_aduana'] = aduana
    else:
        alertas.append("❌ No se encontró número de despacho en el DI.")
        datos['nro_despacho'] = ''
        datos['anio'] = ''
        # Fallback: si no se pudo parsear el número de despacho, intentamos
        # recuperar la aduana buscando su nombre en el texto (mejor esto que nada).
        id_aduana = ''
        for nombre_aduana, codigo_aduana in ADUANAS.items():
            if nombre_aduana in text_norm_upper:
                id_aduana = codigo_aduana
                break
        datos['id_aduana'] = id_aduana

    fechas = re.findall(r'\b(\d{2}/\d{2}/\d{4})\b', text_norm)
    datos['fecha_nac'] = fechas[0] if fechas else ''
    if not fechas:
        alertas.append("❌ No se encontró fecha de oficialización en el DI.")

    cuits = re.findall(r'\b(\d{2}-\d{8}-\d)\b', text_norm)
    if cuits:
        datos['cuit_importador'] = cuits[0]
        datos['cuit_comprador'] = cuits[0]
    else:
        alertas.append("❌ No se encontró CUIT del importador en el DI.")
        datos['cuit_importador'] = ''
        datos['cuit_comprador'] = ''

    datos['cuit_despachante'] = cuits[1] if len(cuits) >= 2 else '20-22824212-9'
    if len(cuits) < 2:
        alertas.append("⚠️ No se encontró CUIT del despachante. Se usará el valor por defecto.")

    m = re.search(r'(FINNING\s+\S+(?:\s+\S+){1,3})', text_norm.upper())
    datos['importador'] = m.group(1).strip() if m else 'FINNING SOLUCIONES MINERAS SA'

    # ─ País de fabricación / procedencia, POR ÍTEM ─
    # FIX: en despachos con varios ítems (ej: el motor + repuestos sueltos
    # en el mismo DI), tomar el primer renglón "Origen País / Procedencia"
    # de TODO el texto puede traer el país de un ítem que no es el motor
    # (ej: tornillos clasificados en otra posición arancelaria). Acá se
    # identifican específicamente los ítems cuya posición SIM empieza con
    # 8408 o 8409 (motores de émbolo diesel/semi-diesel y sus partes, que
    # es la familia arancelaria de motores/blocks) y se extrae el país de
    # CADA uno de esos ítems puntualmente, en el orden en que aparecen.
    # `paises_por_item` queda disponible para que el flujo principal le
    # asigne a cada ENGINE/BLOCK cargado el país de su propio ítem del DI,
    # en vez de un único país "global" para todo el despacho.
    datos['paises_por_item'] = []
    for m_item in re.finditer(r'\d{4}\s+N\s+(840[89]\.\d{2}\.\d{2}\.\d{3}[A-Z]?)', text_norm_upper):
        pos_after = m_item.end()
        m_val = re.search(r'[\d.,]+\s+.+?(UNIDAD|KILOGRAMO)\s', text_norm_upper[pos_after:pos_after + 600])
        if not m_val:
            continue
        val_line = m_val.group(0)
        encontrados = []  # (posicion, codigo)
        for pais, codigo in PAISES.items():
            pos = val_line.find(pais)
            if pos != -1:
                encontrados.append((pos, codigo))
        encontrados.sort(key=lambda x: x[0])
        codigos_ordenados = []
        for _, codigo in encontrados:
            if codigo not in codigos_ordenados:
                codigos_ordenados.append(codigo)
        if len(codigos_ordenados) >= 2:
            datos['paises_por_item'].append({
                'fabricacion': codigos_ordenados[0], 'procedencia': codigos_ordenados[1],
            })
        elif len(codigos_ordenados) == 1:
            datos['paises_por_item'].append({
                'fabricacion': codigos_ordenados[0], 'procedencia': codigos_ordenados[0],
            })

    if datos['paises_por_item']:
        # Compatibilidad hacia atrás: pais_fabricacion/procedencia "global"
        # quedan como el del primer ítem motor encontrado (sigue sirviendo
        # de fallback en despachos de un solo ítem).
        datos['pais_fabricacion'] = datos['paises_por_item'][0]['fabricacion']
        datos['pais_procedencia'] = datos['paises_por_item'][0]['procedencia']
    else:
        # Fallback: ningún ítem con posición 8408/8409 (documento atípico).
        # Se recurre al comportamiento anterior: primer renglón "Origen
        # País/Procedencia" que aparezca en el texto.
        datos['pais_procedencia'] = ''
        datos['pais_fabricacion'] = ''
        lines = text_norm_upper.split('\n')
        for i, line in enumerate(lines):
            if 'ORIGEN' in line and ('PROCEDENCIA' in line or 'PAIS' in line):
                if i + 1 < len(lines):
                    val_line = lines[i + 1].strip()
                    encontrados = []
                    for pais, codigo in PAISES.items():
                        pos = val_line.find(pais)
                        if pos != -1:
                            encontrados.append((pos, codigo))
                    encontrados.sort(key=lambda x: x[0])
                    codigos_ordenados = []
                    for _, codigo in encontrados:
                        if codigo not in codigos_ordenados:
                            codigos_ordenados.append(codigo)
                    if len(codigos_ordenados) >= 2:
                        datos['pais_fabricacion'] = codigos_ordenados[0]
                        datos['pais_procedencia'] = codigos_ordenados[1]
                    elif len(codigos_ordenados) == 1:
                        datos['pais_fabricacion'] = codigos_ordenados[0]
                        datos['pais_procedencia'] = codigos_ordenados[0]
                    break

    if not datos['pais_procedencia']:
        for pais, codigo in PAISES.items():
            if pais in text_norm_upper:
                datos['pais_procedencia'] = codigo
                if not datos['pais_fabricacion']:
                    datos['pais_fabricacion'] = codigo
                break

    if not datos['pais_procedencia']:
        alertas.append("⚠️ No se encontró país de procedencia en el DI.")
    if not datos['pais_fabricacion']:
        alertas.append("⚠️ No se encontró país de fabricación en el DI.")

    datos['regimen'] = '20'

    m = re.search(r'ZA\(0*(\d{4})\)', text_norm)
    datos['anio_fab_di'] = m.group(1) if m else ''

    return datos, alertas


# ─── PARSEO DNRPA ───

# Los códigos de tipo son FIJOS en el sistema DNRPA (no dependen del OCR):
# 09 = BLOCK, 23 = MOTOR.
CODIGOS_TIPO_FIJOS = {'BLOCK': '09', 'MOTOR': '23'}

# Palabras de encabezado que aparecen siempre en la tabla "Consulta
# Marca-Tipo-Modelo" y que NO deben confundirse con la descripción de marca.
_DNRPA_STOPWORDS = {
    'CONSULTA', 'TABLA', 'MARCA', 'TIPO', 'TIPOS', 'MODELO', 'MODELOS',
    'CODIGO', 'CÓDIGO', 'DESCRIPCION', 'DESCRIPCIÓN', 'DENOMINACION',
    'DENOMINACIÓN', 'CERTIFICADO', 'PESO', 'UNIDAD', 'BLOCK', 'MOTOR',
}


def parsear_dnrpa(text, label=""):
    """
    El DNRPA suele ser una captura de pantalla (sin capa de texto), por lo
    que estos datos vienen siempre de OCR. El OCR puede reordenar las
    columnas de la tabla según la resolución/segmentación, así que en vez de
    depender de un único regex secuencial ("165 CATERPILLAR C52 C2.8" en ese
    orden exacto), buscamos cada dato de forma independiente en todo el
    texto. Esto es mucho más tolerante a que el OCR separe las columnas en
    líneas distintas.
    """
    datos = {}
    alertas = []
    text_upper = text.upper()

    # ─ ID de marca: primer número de 3 dígitos "suelto" en la página ─
    m_id = re.search(r'\b(\d{3})\b', text_upper)
    id_marca = m_id.group(1) if m_id else ''

    # ─ Descripción de marca: primera palabra de 4+ letras que no sea
    #   un encabezado de tabla conocido (ej: CATERPILLAR) ─
    palabras = re.findall(r'\b[A-ZÁÉÍÓÚÑ]{4,}\b', text_upper)
    candidatas = [p for p in palabras if p not in _DNRPA_STOPWORDS]
    marca_desc = candidatas[0] if candidatas else ''

    # ─ Modelo: patrones tipo "C52", "C2.8", "C2", etc. ─
    modelos = re.findall(r'\b([A-Z]{1,3}\d+(?:\.\d+)?)\b', text_upper)
    id_modelo = modelos[0] if modelos else ''
    cm_modelo = modelos[1] if len(modelos) > 1 else id_modelo

    if id_marca and marca_desc:
        datos['id_marca'] = id_marca
        datos['marca_desc'] = marca_desc
        datos['id_modelo'] = id_modelo
        datos['cm_modelo'] = cm_modelo
    else:
        alertas.append(f"❌ No se encontró marca/modelo en DNRPA {label}.")
        datos['id_marca'] = id_marca
        datos['id_modelo'] = id_modelo

    # ─ Tipos (BLOCK/MOTOR) ─
    # No dependemos de que el OCR lea correctamente el código numérico (23,
    # 09): esos códigos son fijos, así que solo necesitamos detectar la
    # palabra BLOCK o MOTOR en el texto.
    datos['tipos'] = {}
    for tipo_key, codigo_fijo in CODIGOS_TIPO_FIJOS.items():
        idx_tipo = text_upper.find(tipo_key)
        if idx_tipo == -1:
            continue
        contexto = text_upper[idx_tipo: idx_tipo + 200]
        peso_m = re.search(r'(\d[\d,\.]*)\s*(KGS?|C\.?C\.?)', contexto, re.IGNORECASE)
        peso = peso_m.group(1).replace(',', '').replace('.', '') if peso_m else ''
        datos['tipos'][tipo_key] = {'codigo': codigo_fijo, 'peso': peso}

    if not datos['tipos']:
        alertas.append(f"❌ No se encontraron tipos (BLOCK/MOTOR) en DNRPA {label}.")

    return datos, alertas


def _clean_celda_html(td):
    """Limpia el texto de una celda de la consulta DNRPA (saca &nbsp; y espacios extra)."""
    text = td.get_text()
    text = text.replace('\xa0', ' ')
    return re.sub(r'\s+', ' ', text).strip()


def parsear_dnrpa_html(html_bytes, label=""):
    """
    Parseo del DNRPA cuando el operador sube el .htm/.html de la consulta
    "Marca-Tipo-Modelo" en vez de un PDF/captura. Es el camino preferido:
    el HTML trae el dato como texto real, no como imagen, así que no
    depende del OCR (nada de resoluciones, recortes ni capturas que fallan
    distinto cada vez).

    Estructura esperada (siempre la misma en esta consulta): las tablas de
    encabezado tienen bgcolor="#2B6D85" (celeste) y las tablas con la fila
    de datos real tienen bgcolor="#FFFFFF" (blanco). La primera tabla
    blanca es Marca/Modelo (4 celdas: id_marca, marca_desc, id_modelo,
    cm_modelo); la segunda es Tipos (4 celdas: código, denominación,
    certificado, peso unidad).
    """
    datos = {}
    alertas = []

    try:
        html_text = html_bytes.decode('utf-8')
    except UnicodeDecodeError:
        html_text = html_bytes.decode('latin-1', errors='ignore')

    soup = BeautifulSoup(html_text, 'html.parser')

    filas_blancas = []
    for table in soup.find_all('table'):
        bgcolor = (table.get('bgcolor') or '').upper()
        if bgcolor in ('#FFFFFF', '#FFF', 'WHITE'):
            for fila in table.find_all('tr'):
                celdas = [_clean_celda_html(td) for td in fila.find_all('td')]
                celdas = [c for c in celdas if c]
                if celdas:
                    filas_blancas.append(celdas)

    # ─ Marca / Modelo (primera fila blanca) ─
    if filas_blancas and len(filas_blancas[0]) >= 2:
        fila = filas_blancas[0]
        datos['id_marca'] = fila[0]
        datos['marca_desc'] = fila[1]
        datos['id_modelo'] = fila[2] if len(fila) > 2 else ''
        datos['cm_modelo'] = fila[3] if len(fila) > 3 else datos['id_modelo']
    else:
        alertas.append(f"❌ No se encontró marca/modelo en DNRPA {label} (HTML).")
        datos['id_marca'] = ''
        datos['id_modelo'] = ''

    # ─ Tipos (todas las filas blancas desde la segunda en adelante) ─
    # FIX: algunos motores tienen tanto código BLOCK como código MOTOR
    # (el mismo motor sirve para las dos consultas), entonces la tabla
    # "Tipos" trae DOS filas de datos, no una. Antes solo se leía
    # filas_blancas[1] y se perdía la segunda fila (ej: se guardaba BLOCK
    # y se ignoraba MOTOR, o viceversa). Ahora se recorren todas.
    datos['tipos'] = {}
    for fila in filas_blancas[1:]:
        if len(fila) < 2:
            continue
        codigo = fila[0]
        denominacion = fila[1].upper()
        peso_raw = fila[3] if len(fila) > 3 else (fila[2] if len(fila) > 2 else '')
        peso_m = re.search(r'(\d[\d.,]*)', peso_raw)
        peso = peso_m.group(1).replace('.', '').replace(',', '') if peso_m else ''

        if 'MOTOR' in denominacion:
            tipo_key = 'MOTOR'
        elif 'BLOCK' in denominacion:
            tipo_key = 'BLOCK'
        else:
            tipo_key = None

        if tipo_key:
            datos['tipos'][tipo_key] = {'codigo': codigo, 'peso': peso}

    if not datos['tipos']:
        alertas.append(f"❌ No se encontraron tipos (BLOCK/MOTOR) en DNRPA {label} (HTML).")

    return datos, alertas


# ─── PARSEO FACTURA ───

def parsear_facturas_streaming(fc_files, n_engines):
    motores = []
    for fc_f in fc_files:
        if len(motores) >= n_engines:
            break
        fc_bytes = fc_f.read()
        text_total = extract_text_pdfplumber(fc_bytes)
        if text_total and len(text_total.strip()) > 50:
            for line in text_total.split('\n'):
                uid = re.search(r'UNIQUE\s+ID[:\s]+([A-Z0-9]+)', line, re.IGNORECASE)
                if uid and uid.group(1) not in motores:
                    motores.append(uid.group(1))
        else:
            try:
                with pdfplumber.open(BytesIO(fc_bytes)) as pdf:
                    total_pages = len(pdf.pages)
            except:
                total_pages = 0
            for page_num in range(total_pages):
                if len(motores) >= n_engines:
                    break
                page_text = ocr_pdf_bytes(fc_bytes, f"fc_p{page_num}", dpi=200)
                for line in page_text.split('\n'):
                    uid = re.search(r'UNIQUE\s+ID[:\s]+([A-Z0-9]+)', line, re.IGNORECASE)
                    if uid and uid.group(1) not in motores:
                        motores.append(uid.group(1))
    return motores


# ─── GENERAR TXT ───

def generar_txt(di, items_procesados, lcm_valor):
    try:
        fecha_dt = datetime.datetime.strptime(di['fecha_nac'], "%d/%m/%Y")
        anio_dos = str(fecha_dt.year)[-2:]
        fecha_str = fecha_dt.strftime("%d/%m/%Y")
    except:
        anio_dos = di.get('anio', '26')
        fecha_str = di.get('fecha_nac', '')

    nro_despacho = f"{di['nro_despacho']}/{anio_dos}"
    id_aduana = di.get('id_aduana', '001')

    if lcm_valor and lcm_valor.strip():
        parts = (re.split(r'[/\-\s]+', lcm_valor.strip()) + ["0","0","0"])[:3]
        lcm_tipo, lcm_nro, lcm_anio = parts
    else:
        lcm_tipo, lcm_nro, lcm_anio = "0", "0", "0"

    def q(v): return f'"{v}"'
    def safe(v): return str(v).strip().replace(" ", "") if v else ""

    caratula = ";".join([
        q(id_aduana), q(nro_despacho), q("00"), q("12"),
        q(di.get('cuit_importador','')), q("12"),
        q(di.get('cuit_comprador','')), q("12"),
        q(di.get('cuit_despachante','')), q(di.get('regimen','20')),
        q(fecha_str), q(di.get('pais_procedencia','212')),
        q(str(len(items_procesados))), q("N"), q("S"),
        q(""), q(""), q(""), q(""), q("")
    ])

    lineas = []
    for i, item in enumerate(items_procesados, start=1):
        dnrpa = item['dnrpa']
        tipo = item['tipo']
        tipo_key = 'MOTOR' if tipo == 'ENGINE' else 'BLOCK'
        id_tipo = dnrpa.get('tipos',{}).get(tipo_key,{}).get('codigo','')
        peso = dnrpa.get('tipos',{}).get(tipo_key,{}).get('peso','')
        nro_motor = safe(item.get('motor','')) if tipo == 'ENGINE' else ''
        anio = str(item['anio_fab'])
        linea = ";".join([
            q(id_aduana), q(nro_despacho), q("00"), q(str(i)),
            q(dnrpa.get('id_marca','')), q(id_tipo), q(dnrpa.get('id_modelo','')),
            q(lcm_tipo), q(lcm_nro), q(lcm_anio),
            q(anio), q(anio),
            q(dnrpa.get('id_marca','')), q(nro_motor),
            q("000"), q("NOPOSEE"),
            q(item.get('pais_fabricacion', di.get('pais_fabricacion', di.get('pais_procedencia','212')))),
            q(str(peso)), q("N")
        ])
        lineas.append(linea)

    return caratula + "\n" + "\n".join(lineas)


# ─── GENERAR EXCEL ───

def generar_excel(di, items_procesados, lcm_valor):
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb['ANVERSO']

    try:
        fecha_dt = datetime.datetime.strptime(di['fecha_nac'], "%d/%m/%Y")
    except:
        fecha_dt = datetime.datetime.now()

    ws['E3'] = di['nro_despacho']
    ws['J3'] = fecha_dt
    ws['L3'] = di.get('regimen', '20')
    ws['E7'] = di.get('importador', '')
    ws['L7'] = di.get('cuit_importador', '')
    ws['I9'] = di.get('importador', '')
    ws['L9'] = di.get('cuit_comprador', '')
    try: ws['E11'] = int(di.get('pais_procedencia', 212))
    except: ws['E11'] = di.get('pais_procedencia', 212)

    for row_idx in range(16, 31):
        for col_idx in range(1, 14):
            ws.cell(row=row_idx, column=col_idx).value = None

    lcm_excel = lcm_valor.strip() if lcm_valor and lcm_valor.strip() else 'XXX'

    for i, item in enumerate(items_procesados):
        row = 16 + i
        dnrpa = item['dnrpa']
        tipo = item['tipo']
        tipo_key = 'MOTOR' if tipo == 'ENGINE' else 'BLOCK'
        id_tipo = dnrpa.get('tipos',{}).get(tipo_key,{}).get('codigo','')
        peso = dnrpa.get('tipos',{}).get(tipo_key,{}).get('peso','')
        nro_motor = item.get('motor','') if tipo == 'ENGINE' else ''
        anio = str(item['anio_fab'])

        ws.cell(row=row, column=1).value = i + 1
        ws.cell(row=row, column=2).value = dnrpa.get('id_marca','')
        ws.cell(row=row, column=3).value = id_tipo
        ws.cell(row=row, column=4).value = dnrpa.get('id_modelo','')
        ws.cell(row=row, column=5).value = lcm_excel
        ws.cell(row=row, column=6).value = anio
        ws.cell(row=row, column=7).value = anio
        ws.cell(row=row, column=8).value = dnrpa.get('id_marca','')
        ws.cell(row=row, column=9).value = nro_motor
        ws.cell(row=row, column=10).value = '000'
        ws.cell(row=row, column=11).value = 'NO POSEE'
        ws.cell(row=row, column=12).value = item.get('pais_fabricacion', di.get('pais_fabricacion', di.get('pais_procedencia','212')))
        ws.cell(row=row, column=13).value = str(peso)

    ws['D35'] = 'CAPITAL FEDERAL'
    ws['E37'] = datetime.datetime.now()

    ADUANAS_NOMBRE = {
        '001': '001-BS.AS. CAPITAL', '003': '003-BAHIA BLANCA', '004': '004-BARILOCHE',
        '008': '008-CAMPANA', '017': '017-CORDOBA', '029': '029-IGUAZU',
        '033': '033-LA PLATA', '037': '037-MAR DEL PLATA', '038': '038-MENDOZA',
        '052': '052-ROSARIO', '053': '053-SALTA', '055': '055-SAN JUAN',
        '073': '073-EZEIZA', '074': '074-TUCUMAN', '075': '075-NEUQUEN',
        '091': '091-BS.AS. NORTE', '092': '092-BS.AS. SUR',
    }
    id_aduana = di.get('id_aduana', '001')
    ws['D31'] = ADUANAS_NOMBRE.get(id_aduana, f"{id_aduana}-")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════
# INTERFAZ
# ═══════════════════════════════════════════════

st.markdown('<p class="section-title">1 · Documentos generales</p>', unsafe_allow_html=True)
col1, col2 = st.columns(2)
with col1:
    di_file = st.file_uploader("📋 DI (PDF)", type="pdf")
with col2:
    fc_files = st.file_uploader("🧾 Factura/s (PDF)", type="pdf", accept_multiple_files=True)

st.markdown('<p class="section-title">2 · Ítems de la DJIM</p>', unsafe_allow_html=True)
st.caption("Agregá un ítem por cada motor o block del despacho.")

col_add, col_rem = st.columns([1, 1])
with col_add:
    if st.button("➕ Agregar ítem"):
        st.session_state.n_items += 1
with col_rem:
    if st.session_state.n_items > 0:
        if st.button("➖ Quitar último"):
            st.session_state.n_items -= 1

tipos_seleccionados = []
dnrpa_files = []
anios_block = []

for idx in range(st.session_state.n_items):
    st.markdown(f"**Ítem {idx+1}**")
    col1, col2 = st.columns([1, 2])
    with col1:
        tipo = st.selectbox("Tipo", ["ENGINE", "BLOCK"], key=f"tipo_sel_{idx}")
        tipos_seleccionados.append(tipo)
        if tipo == "BLOCK":
            anio = st.text_input("Año fabricación", key=f"anio_sel_{idx}", placeholder="ej: 2025")
            anios_block.append(anio)
        else:
            anios_block.append("")
    with col2:
        dnrpa = st.file_uploader(
            "DNRPA (.htm recomendado, o PDF)",
            type=["htm", "html", "pdf"],
            key=f"dnrpa_sel_{idx}",
        )
        dnrpa_files.append(dnrpa)
    st.divider()

st.markdown('<p class="section-title">3 · Datos adicionales</p>', unsafe_allow_html=True)
col1, col2 = st.columns(2)
with col1:
    tiene_lcm = st.radio("¿Tiene LCM?", ["No", "Sí"], horizontal=True)
with col2:
    lcm_valor = ""
    if tiene_lcm == "Sí":
        lcm_valor = st.text_input("Número LCM", placeholder="ej: 39/12345/2025")

st.markdown("---")

if st.button("⚙️ Procesar y Generar", type="primary", use_container_width=True):

    errores = []
    if not di_file:
        errores.append("❌ Faltá subir el DI.")
    if not fc_files:
        errores.append("❌ Faltá subir al menos una factura.")
    if st.session_state.n_items == 0:
        errores.append("❌ Agregá al menos un ítem.")
    for idx in range(st.session_state.n_items):
        if not dnrpa_files[idx]:
            errores.append(f"❌ Faltá el DNRPA del ítem {idx+1}.")
        if tipos_seleccionados[idx] == 'BLOCK' and not anios_block[idx].strip():
            errores.append(f"❌ Ingresá el año de fabricación del ítem {idx+1} (BLOCK).")

    if errores:
        for e in errores:
            st.error(e)
        st.stop()

    with st.spinner("Procesando documentos..."):
        di_bytes = di_file.read()
        di_text = get_text_di(di_bytes, "di", dpi=150)
        di_datos, di_alertas = parsear_di(di_text)

        n_engines = sum(1 for t in tipos_seleccionados if t == 'ENGINE')
        motores_factura = parsear_facturas_streaming(fc_files, n_engines)

        items_procesados = []
        todas_alertas = di_alertas.copy()
        motor_idx = 0
        paises_item_idx = 0

        for idx in range(st.session_state.n_items):
            tipo = tipos_seleccionados[idx]
            tipo_key = 'MOTOR' if tipo == 'ENGINE' else 'BLOCK'

            dnrpa_bytes = dnrpa_files[idx].read()
            dnrpa_nombre = (dnrpa_files[idx].name or "").lower()

            if dnrpa_nombre.endswith(".htm") or dnrpa_nombre.endswith(".html"):
                # Camino preferido: el HTML de la consulta DNRPA trae el dato
                # como texto real, sin depender de OCR.
                dnrpa_datos, dnrpa_alertas = parsear_dnrpa_html(dnrpa_bytes, f"ítem {idx+1}")
            else:
                # Fallback para capturas/PDF viejos: psm=6 + upscale=3 porque
                # el DNRPA en PDF suele ser una captura de pantalla de baja
                # resolución; agrandamos la imagen y forzamos modo de
                # segmentación de bloque uniforme para mejorar el OCR.
                dnrpa_text = get_text(dnrpa_bytes, f"dnrpa_{idx}", dpi=250, psm=6, upscale=3)
                dnrpa_datos, dnrpa_alertas = parsear_dnrpa(dnrpa_text, f"ítem {idx+1}")

            todas_alertas.extend(dnrpa_alertas)

            if tipo == 'ENGINE':
                anio_fab = di_datos.get('anio_fab_di', '')
                if not anio_fab:
                    todas_alertas.append(f"❌ No se encontró año de fabricación en el DI para ENGINE ítem {idx+1}.")
            else:
                anio_fab = anios_block[idx]

            motor = ''
            if tipo == 'ENGINE':
                if motor_idx < len(motores_factura):
                    motor = motores_factura[motor_idx]
                    motor_idx += 1
                else:
                    todas_alertas.append(f"⚠️ No se encontró UNIQUE ID para ENGINE ítem {idx+1}.")

            if not dnrpa_datos.get('tipos',{}).get(tipo_key,{}).get('peso'):
                todas_alertas.append(f"❌ No se encontró peso para {tipo} en DNRPA ítem {idx+1}.")

            # País de fabricación/procedencia de ESTE ítem puntual: si el DI
            # tiene varios ítems con posición 8408/8409 (motor + repuestos
            # en el mismo despacho), cada ENGINE/BLOCK cargado toma el país
            # del ítem del DI que le corresponde en orden, no un país
            # "global" único para todo el despacho.
            paises_item = di_datos.get('paises_por_item') or []
            if paises_item_idx < len(paises_item):
                pais_fab_item = paises_item[paises_item_idx]['fabricacion']
                paises_item_idx += 1
            else:
                pais_fab_item = di_datos.get('pais_fabricacion', di_datos.get('pais_procedencia', '212'))

            items_procesados.append({
                'tipo': tipo, 'dnrpa': dnrpa_datos,
                'anio_fab': anio_fab, 'motor': motor,
                'pais_fabricacion': pais_fab_item,
            })

        if n_engines > len(motores_factura):
            todas_alertas.append(
                f"⚠️ Se declararon {n_engines} ENGINE(s) pero se encontraron "
                f"solo {len(motores_factura)} UNIQUE ID(s). Verificar manualmente."
            )

        st.session_state['resultado_txt'] = generar_txt(di_datos, items_procesados, lcm_valor)
        if os.path.exists(TEMPLATE_PATH):
            excel_buf = generar_excel(di_datos, items_procesados, lcm_valor)
            st.session_state['resultado_excel'] = excel_buf.read()
            st.session_state['resultado_nro'] = di_datos.get('nro_despacho', 'DJIM')

    for a in [x for x in todas_alertas if x.startswith("⚠️")]:
        st.warning(a)

    errores_criticos = [x for x in todas_alertas if x.startswith("❌")]
    if errores_criticos:
        for e in errores_criticos:
            st.error(e)
        st.stop()

    st.markdown('<div class="alerta-ok">✅ Documentos procesados correctamente.</div>', unsafe_allow_html=True)
    st.markdown("")

    with st.expander("📋 Ver datos extraídos"):
        st.markdown("**DI:**")
        st.json({k: v for k, v in di_datos.items() if k != 'anio_fab_di'})
        for idx, item in enumerate(items_procesados):
            st.markdown(f"**Ítem {idx+1} — {item['tipo']}:**")
            st.json({
                'id_marca': item['dnrpa'].get('id_marca'),
                'id_modelo': item['dnrpa'].get('id_modelo'),
                'tipos': item['dnrpa'].get('tipos'),
                'anio_fab': item['anio_fab'],
                'motor': item.get('motor',''),
            })

# ── DESCARGAS PERSISTENTES ──
if 'resultado_txt' in st.session_state or 'resultado_excel' in st.session_state:
    st.markdown('<p class="section-title">4 · Descargar</p>', unsafe_allow_html=True)
    col1, col2 = st.columns(2)
    with col1:
        if 'resultado_txt' in st.session_state:
            st.download_button(
                "📥 DJIM Electrónica (.txt)",
                data=st.session_state['resultado_txt'].encode('utf-8'),
                file_name="DJIM_ELECTRONICA.txt",
                mime="text/plain",
                use_container_width=True,
                key="dl_txt"
            )
    with col2:
        if 'resultado_excel' in st.session_state:
            nro = st.session_state.get('resultado_nro', 'DJIM')
            st.download_button(
                "📥 DJIM Excel (.xlsx)",
                data=st.session_state['resultado_excel'],
                file_name=f"DJIM_{nro}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="dl_excel"
            )
