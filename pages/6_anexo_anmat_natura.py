import streamlit as st
import pandas as pd
import numpy as np
import unicodedata
import re
import os
import subprocess
import tempfile
import zipfile
from io import BytesIO
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

st.markdown("""
<style>
    .header-box {
        background: linear-gradient(135deg, #e8f4f8 0%, #d0eaf5 100%);
        border-left: 5px solid #00b4d8;
        border-radius: 8px;
        padding: 20px 28px;
        margin-bottom: 24px;
    }
    .header-box h1 { color: #00b4d8; font-size: 1.6rem; font-weight: 700; margin: 0 0 4px 0; }
    .header-box p { color: #444; font-size: 0.9rem; margin: 0; }
    .card { background: #ffffff; border: 1px solid #dde3ea; border-radius: 8px; padding: 20px; margin-bottom: 16px; box-shadow: 0 1px 4px rgba(0,0,0,0.06); }
    .card h3 { color: #00b4d8; font-size: 0.9rem; font-weight: 700; margin: 0 0 14px 0; text-transform: uppercase; letter-spacing: 0.06em; }
    .alert-box { background: #fff5f5; border: 1px solid #ffb3b3; border-radius: 6px; padding: 12px 16px; margin: 6px 0; color: #cc0000; font-size: 0.9rem; }
    .alert-box strong { color: #990000; }
    .success-box { background: #f0fff8; border: 1px solid #00c896; border-radius: 6px; padding: 12px 16px; margin: 6px 0; color: #007a5c; font-size: 0.9rem; }
    .info-box { background: #f0f8ff; border: 1px solid #90cce8; border-radius: 6px; padding: 12px 16px; margin: 6px 0; color: #0066aa; font-size: 0.9rem; }
    .stat-card { background: #ffffff; border: 1px solid #dde3ea; border-radius: 8px; padding: 16px; text-align: center; box-shadow: 0 1px 4px rgba(0,0,0,0.05); }
    .stat-card .number { color: #00b4d8; font-size: 2rem; font-weight: 700; line-height: 1; }
    .stat-card .label { color: #888; font-size: 0.78rem; margin-top: 4px; text-transform: uppercase; letter-spacing: 0.05em; }
    .step-badge { display: inline-block; background: #00b4d8; color: #ffffff; border-radius: 50%; width: 22px; height: 22px; text-align: center; line-height: 22px; font-weight: 700; font-size: 0.78rem; margin-right: 8px; }
    .modo-muestras { background: linear-gradient(135deg, #fff8e1 0%, #fff3cd 100%); border-left: 5px solid #ffc107; border-radius: 8px; padding: 12px 18px; margin-bottom: 16px; color: #856404; font-weight: 600; font-size: 0.95rem; }
    .equiv-found { background: #f0fff8; border: 1px solid #00c896; border-radius: 6px; padding: 10px 14px; margin: 6px 0; color: #007a5c; font-size: 0.85rem; }
    .equiv-notfound { background: #fff5f5; border: 1px solid #ffb3b3; border-radius: 6px; padding: 8px 14px; margin: 4px 0; color: #cc0000; font-size: 0.82rem; }
    [data-testid="stToolbar"] { visibility: hidden !important; }
    [data-testid="stDecoration"] { display: none !important; }
    a[href*="github.com"] { display: none !important; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="header-box">
    <h1>📋 Generador de Anexo ANMAT</h1>
    <p>Natura · Avon · Operaciones de importación</p>
</div>
""", unsafe_allow_html=True)

def limpiar_str(s):
    s = str(s).strip()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = s.lower()
    s = s.replace(':', ' ').replace('.', ' ')
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def _codigo_str(v):
    if pd.isna(v):
        return ''
    s = str(v).strip()
    if re.match(r'^\d+\.0$', s):
        s = s[:-2]
    return s

def normalizar_pais(origen_str):
    if pd.isna(origen_str):
        return ''
    s = str(origen_str).strip()
    if ':' in s:
        return s.split(':')[0].strip().lower()
    return s.split(' ')[0].strip().lower()

@st.cache_data
def cargar_anmat(file_bytes):
    buf = BytesIO(file_bytes)
    try:
        df = pd.read_excel(buf, sheet_name='HISTORICO', header=0, engine='pyxlsb')
    except:
        buf.seek(0)
        df = pd.read_excel(buf, sheet_name='HISTORICO', header=0)
    df['CM'] = df['CM'].apply(_codigo_str)

    def _normalizar_col(df, col_estandar, variantes):
        if col_estandar in df.columns:
            return df
        for v in variantes:
            if v in df.columns:
                return df.rename(columns={v: col_estandar})
        cols_upper = {c.upper(): c for c in df.columns}
        for v in variantes:
            if v.upper() in cols_upper:
                return df.rename(columns={cols_upper[v.upper()]: col_estandar})
        return df

    df = _normalizar_col(df, 'NOMBRE', ['DESCRIPCION', 'DESCRIPTION', 'NOMBRE DEL PRODUCTO', 'PRODUCTO', 'NOMBRE DE REGISTRO DE PRODUCTO', 'NOMBRE REGISTRO DE PRODUCTO'])
    df = _normalizar_col(df, 'Variedad', ['VARIEDADES', 'Variedades', 'VARIEDAD'])
    df = _normalizar_col(df, 'Registros ANMAT', ['Registro ANMAT', 'REGISTROS ANMAT', 'REGISTRO ANMAT', 'Registros', 'REGISTRO', 'REGISTRO (TRAMITE #)', 'REGISTRO (TRÁMITE #)', 'TRAMITE', 'TRÁMITE'])
    df = _normalizar_col(df, 'ORIGEN', ['ORIGEN/ELABORADOR', 'ORIGEN / ELABORADOR', 'ELABORADOR', 'PAIS DE ORIGEN'])
    df = _normalizar_col(df, 'Fecha Admision', ['FECHA ADMISIÓN', 'FECHA ADMISION', 'FECHA DE ADMISION', 'FECHA DE ADMISIÓN'])
    return df

@st.cache_data
def cargar_avon(file_bytes):
    buf = BytesIO(file_bytes)
    return pd.read_excel(buf, header=0)

@st.cache_data
def cargar_fabricantes(file_bytes, suffix='.xlsx'):
    buf = BytesIO(file_bytes)
    if suffix == '.xls':
        df = pd.read_excel(buf, header=1, engine='xlrd')
    else:
        df = pd.read_excel(buf, header=1)
    df.columns = ['material', 'En Historico', 'Corresponde']
    return df

@st.cache_data
def cargar_ncm(file_bytes):
    buf = BytesIO(file_bytes)
    df = pd.read_excel(buf, header=0)
    df['Artículo'] = df['Artículo'].apply(_codigo_str)
    return df

def cargar_pl(file_bytes):
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        f.write(file_bytes)
        tmp = f.name
    xl = pd.ExcelFile(tmp)
    rows = []
    invoice = None
    for sh in xl.sheet_names:
        df = pd.read_excel(tmp, sheet_name=sh, header=None)
        header_row = None
        for i, row in df.iterrows():
            row_str_raw = ' '.join(str(v).replace('\n', ' ').upper() for v in row.values if pd.notna(v))
            row_str = ''.join(c for c in unicodedata.normalize('NFD', row_str_raw) if unicodedata.category(c) != 'Mn')
            vals_upper = [str(v).replace('\n', ' ').strip().upper() for v in row.values if pd.notna(v)]
            if ('MATERIAL CODE' in row_str or 'MATERIAL\nCODE' in str(row.values)
                    or ('CODE' in vals_upper and any(k in row_str for k in ['PRODUCT', 'LOT', 'DESCRIPTION', 'DESCRIPCION', 'PACKING']))
                    or ('CODIGO' in row_str and any(k in row_str for k in ['DESCRIPCION', 'LOTE', 'CANTIDAD']))
                    or ('MATERIAL' in vals_upper and any(k in row_str for k in ['DESCRIPTION', 'LOT', 'QUANTITY']))
                    or ('CUSTOMER CODE' in row_str and any(k in row_str for k in ['DESCRIPTION', 'BATCH', 'PURCHASE']))
                    or ('CODIGO FIABILA' in row_str)):
                header_row = i
                break
        if header_row is None:
            continue
        if not invoice:
            for i, row in df.iterrows():
                vals = list(row.values)
                for j, val in enumerate(vals):
                    val_str = str(val).replace('\n', ' ')
                    if 'Nº INVOICE:' in val_str or 'N° INVOICE:' in val_str:
                        parte = val_str.split(':', 1)[1].strip()
                        if parte and parte != 'nan':
                            invoice = parte
                            break
                        for k in range(j + 1, len(vals)):
                            v = str(vals[k]).strip()
                            if v and v != 'nan':
                                invoice = v
                                break
                        break
                if invoice:
                    break
        data_start = header_row + 1
        if data_start < len(df):
            primera = df.iloc[data_start]
            tiene_numeros = any(str(v).strip().isdigit() or (len(str(v).strip()) >= 5 and str(v).strip().replace('.','').isdigit()) for v in primera.values if pd.notna(v))
            if not tiene_numeros:
                data_start += 1

        def _sin_acentos(s):
            return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')

        header_vals = [_sin_acentos(str(v).replace('\n', ' ').strip().upper()) if pd.notna(v) else '' for v in df.iloc[header_row].values]
        data = df.iloc[data_start:].copy().reset_index(drop=True)
        data.columns = range(len(data.columns))

        col_mat = None
        for idx, h in enumerate(header_vals):
            if h in ('CODE', 'CODIGO', 'MATERIAL CODE', 'CUSTOMER CODE'):
                col_mat = idx
                break
            if 'MATERIAL CODE' in h or h == 'MATERIAL\nCODE' or 'CUSTOMER CODE' in h:
                col_mat = idx
                break

        if col_mat is None:
            for col_idx in range(min(5, len(data.columns))):
                muestra = data[col_idx].dropna().astype(str)
                if (muestra.str.match(r'^\d{5,}').sum() > 0 or muestra.str.match(r'^\d+-\d{4,}').sum() > 0):
                    col_mat = col_idx
                    break
            if col_mat is None:
                col_mat = 1

        data = data[data[col_mat].astype(str).str.match(r'^\d{5,}$|^\d+-\d{4,}$')]

        col_qty_idx, col_desc_idx, col_lote_idx, col_fecha_idx = 2, 3, 5, 6
        col_origen_idx = None
        lote_encontrado = False
        qty_encontrado = False

        for idx, h in enumerate(header_vals):
            if idx in [0, 1]:
                continue
            if not qty_encontrado and any(k in h for k in ['QUANTITY PC', 'QUANTITY', 'CANTIDAD', 'PCS', 'TOTAL NET WEIGHT', 'NET WEIGHT', 'TOTAL UNIDADES', 'TOTAL UNIDAD']) and 'BOX' not in h and ('TOTAL' not in h or 'WEIGHT' in h or 'UNIDAD' in h):
                col_qty_idx = idx
                qty_encontrado = True
            if any(k in h for k in ['DESCRIPTION', 'DESCRIP']):
                col_desc_idx = idx
            if 'LOT PRODUCT' in h and not lote_encontrado:
                col_lote_idx = idx
                lote_encontrado = True
            elif any(k in h for k in ['LOT NUMBER', 'LOT', 'LOTE', 'BATCH']) and 'SUPPLIER' not in h and 'BOX' not in h and not lote_encontrado:
                col_lote_idx = idx
                lote_encontrado = True
            if any(k in h for k in ['EXPIRE', 'VENC', 'EXPIR', 'EXPIRATION']):
                col_fecha_idx = idx
            if any(k in h for k in ['PAIS DE ORIGEN', 'PAÍS DE ORIGEN', 'COUNTRY OF ORIGIN']):
                col_origen_idx = idx

        if col_fecha_idx in data.columns:
            def normalizar_fecha(v):
                if pd.isna(v) or str(v).strip() in ('', 'nan'):
                    return ''
                if isinstance(v, datetime):
                    return f"{v.month:02d}/{v.year}"
                s = str(v).strip()
                m = re.match(r'(\d{4})-(\d{2})-(\d{2})', s)
                if m:
                    return f"{m.group(2)}/{m.group(1)}"
                return s
            data[col_fecha_idx] = data[col_fecha_idx].apply(normalizar_fecha)

        col_map_reorder = {1: col_mat, 2: col_qty_idx, 3: col_desc_idx, 5: col_lote_idx, 6: col_fecha_idx}
        if any(v != k for k, v in col_map_reorder.items()) and len(data) > 0:
            rename_map = {}
            for std_pos, src_pos in col_map_reorder.items():
                if src_pos in data.columns and src_pos != std_pos:
                    rename_map[src_pos] = f'_col{std_pos}'
            data = data.rename(columns=rename_map)
            for std_pos, src_pos in col_map_reorder.items():
                key = f'_col{std_pos}'
                if key in data.columns:
                    data[std_pos] = data[key]
                elif std_pos not in data.columns:
                    data[std_pos] = ''

        rows.append(data)
    pl = pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()
    return pl, invoice

PAISES_CONOCIDOS = ['brazil', 'brasil', 'colombia', 'china', 'argentina', 'mexico', 'méxico', 'peru', 'perú', 'chile', 'uruguay', 'paraguay', 'bolivia', 'ecuador', 'usa', 'united states', 'estados unidos', 'france', 'francia', 'germany', 'alemania', 'italy', 'italia', 'spain', 'españa', 'japan', 'japón', 'japon', 'korea', 'corea', 'india', 'taiwan']
PAIS_NORMALIZADO = {'brazil': 'Brasil', 'brasil': 'Brasil', 'colombia': 'Colombia', 'china': 'China', 'argentina': 'Argentina', 'mexico': 'México', 'méxico': 'México', 'peru': 'Perú', 'perú': 'Perú', 'chile': 'Chile', 'uruguay': 'Uruguay', 'paraguay': 'Paraguay', 'bolivia': 'Bolivia', 'ecuador': 'Ecuador', 'usa': 'USA', 'united states': 'USA', 'estados unidos': 'USA', 'france': 'Francia', 'francia': 'Francia', 'germany': 'Alemania', 'alemania': 'Alemania', 'italy': 'Italia', 'italia': 'Italia', 'spain': 'España', 'españa': 'España', 'japan': 'Japón', 'japón': 'Japón', 'japon': 'Japón', 'korea': 'Korea', 'corea': 'Korea', 'india': 'India', 'taiwan': 'Taiwan'}

def _extraer_pais_de_texto(texto):
    t = texto.lower()
    for p in PAISES_CONOCIDOS:
        if p in t:
            return PAIS_NORMALIZADO.get(p, p.capitalize())
    return None

def _parsear_pdf_proximas(file_bytes):
    try:
        import pdfplumber
    except ImportError:
        return None, False, None
    texto_completo = ''
    tablas = []
    with pdfplumber.open(BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            texto_completo += (page.extract_text() or '') + '\n'
            t = page.extract_table()
            if t:
                tablas.append(t)
    patron_origen = [r'country\s+of\s+origin\s*[:\-]\s*([A-Za-z\s]+)', r'pa[ií]s\s+de\s+origen\s*[:\-]\s*([A-Za-z\s]+)', r'origen\s*[:\-]\s*([A-Za-z\s]+)', r'origin\s*[:\-]\s*([A-Za-z\s]+)']
    origen_explicito = None
    for pat in patron_origen:
        m = re.search(pat, texto_completo, re.IGNORECASE)
        if m:
            candidato = m.group(1).strip().split('\n')[0].strip()
            pais = _extraer_pais_de_texto(candidato)
            if pais:
                origen_explicito = pais
                break
    origen_proveedor = None
    patron_exportador = [r'exporter\s*/\s*shipper\s*[:\-]?\s*(.+)', r'exporter\s*[:\-]\s*(.+)', r'shipper\s*[:\-]\s*(.+)', r'exportador\s*[:\-]\s*(.+)', r'proveedor\s*[:\-]\s*(.+)']
    for pat in patron_exportador:
        m = re.search(pat, texto_completo, re.IGNORECASE)
        if m:
            linea = m.group(1).strip()
            idx = texto_completo.lower().find(linea.lower())
            bloque = texto_completo[idx:idx+300] if idx >= 0 else linea
            pais = _extraer_pais_de_texto(bloque)
            if pais:
                origen_proveedor = pais
                break
    if not origen_proveedor:
        primeras_lineas = '\n'.join(texto_completo.split('\n')[:15])
        origen_proveedor = _extraer_pais_de_texto(primeras_lineas)
    col_material_keywords = ['code', 'código', 'codigo', 'material', 'material code', 'item', 'article', 'artículo', 'articulo', 'ref', 'sku']
    registros = []
    origen_tabla = origen_explicito
    for tabla in tablas:
        if not tabla or len(tabla) < 2:
            continue
        headers = [str(h).strip().lower() if h else '' for h in tabla[0]]
        col_mat_idx = None
        for i, h in enumerate(headers):
            if any(kw in h for kw in col_material_keywords):
                col_mat_idx = i
                break
        col_orig_idx = None
        for i, h in enumerate(headers):
            if any(kw in h for kw in ['origen', 'origin', 'país', 'pais', 'country']):
                col_orig_idx = i
                break
        if col_mat_idx is None:
            continue
        for row in tabla[1:]:
            if not row or col_mat_idx >= len(row):
                continue
            val = str(row[col_mat_idx]).strip() if row[col_mat_idx] else ''
            if re.match(r'^\d{5,}', val):
                orig = origen_tabla
                if col_orig_idx is not None and col_orig_idx < len(row):
                    orig_cell = str(row[col_orig_idx]).strip()
                    p = _extraer_pais_de_texto(orig_cell)
                    if p:
                        orig = p
                registros.append({'Material': val, 'Origen': orig or ''})
    if not registros:
        codigos = re.findall(r'(\d{5,8})', texto_completo)
        for cod in set(codigos):
            registros.append({'Material': cod, 'Origen': origen_tabla or ''})
    if not registros:
        return None, origen_explicito is not None, origen_proveedor
    df = pd.DataFrame(registros)
    df['Material'] = df['Material'].astype(str).str.strip()
    return df, origen_explicito is not None, origen_proveedor

def cargar_proximas(file_bytes, filename=''):
    es_pdf = filename.lower().endswith('.pdf')
    if es_pdf:
        df, origen_explicito, origen_proveedor = _parsear_pdf_proximas(file_bytes)
        return df, es_pdf, origen_explicito, origen_proveedor
    else:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            f.write(file_bytes)
            tmp = f.name
        df = pd.read_excel(tmp, header=0)
        col_map = {c.strip().lower(): c for c in df.columns}
        col_mat = None
        for posible in ['material', 'código de sku', 'codigo de sku', 'sku', 'cod. material', 'codigo material', 'código material', 'article', 'artículo', 'articulo', 'material code', 'item', 'código', 'codigo']:
            if posible in col_map:
                col_mat = col_map[posible]
                break
        if col_mat and col_mat != 'Material':
            df = df.rename(columns={col_mat: 'Material'})
        elif 'Material' not in df.columns:
            for col in df.columns:
                if df[col].dropna().astype(str).str.match(r'^\d{5,}').any():
                    df = df.rename(columns={col: 'Material'})
                    break
        df['Material'] = df['Material'].apply(_codigo_str)
        return df, False, True, None

def buscar_anmat(mat_code, df_anmat):
    found = df_anmat[df_anmat['CM'] == str(mat_code)]
    if len(found) == 0:
        return None
    if len(found) > 1:
        found = found.sort_values('Fecha Admision', ascending=False)
    return found.iloc[0]

def _col_avon(df, variantes):
    cols_norm = {c.strip().lower().replace(' ', '').replace('/', ''): c for c in df.columns}
    for v in variantes:
        key = v.strip().lower().replace(' ', '').replace('/', '')
        if key in cols_norm:
            return cols_norm[key]
    return None

def buscar_avon(mat_code, df_avon):
    mat_str = str(mat_code).strip()
    col_cm = _col_avon(df_avon, ['CM / ZPAC', 'CM/ZPAC', 'CM/ ZPAC', 'CM /ZPAC'])
    col_fi = _col_avon(df_avon, ['FI Code Local', 'FI Code', 'FICodeLocal'])
    if col_cm:
        found = df_avon[df_avon[col_cm].apply(_codigo_str) == mat_str]
        if len(found) > 0:
            return found.iloc[0]
    if col_fi:
        found = df_avon[df_avon[col_fi].apply(_codigo_str) == mat_str]
        if len(found) > 0:
            return found.iloc[0]
    return None

def buscar_fabricante(origen_str, mat_code, df_fab):
    origen = str(origen_str).strip()
    origen_limpio = limpiar_str(origen)
    match_norm = match_parcial = match_row_norm = match_row_parcial = None
    for _, row in df_fab.iterrows():
        en_hist = str(row['En Historico']).strip()
        if not en_hist or en_hist == 'nan':
            continue
        en_hist_limpio = limpiar_str(en_hist)
        if origen == en_hist:
            return row['Corresponde'], None, row
        if match_norm is None and origen_limpio == en_hist_limpio:
            match_norm = row['Corresponde']; match_row_norm = row
        if match_parcial is None and en_hist_limpio in origen_limpio:
            match_parcial = row['Corresponde']; match_row_parcial = row
    if match_norm:
        return match_norm, None, match_row_norm
    if match_parcial:
        return match_parcial, None, match_row_parcial
    return None, f"Fabricante no encontrado para material {mat_code} (origen: {origen_str})", None

def extraer_origen_de_fila_fab(fab_row):
    if fab_row is None:
        return None, []
    textos = [str(fab_row.get('En Historico', '')), str(fab_row.get('Corresponde', ''))]
    paises_encontrados = []
    for texto in textos:
        p = _extraer_pais_de_texto(texto)
        if p and p not in paises_encontrados:
            paises_encontrados.append(p)
    if len(paises_encontrados) == 1:
        return paises_encontrados[0], paises_encontrados
    return None, paises_encontrados

def buscar_ncm(mat_code, df_ncm):
    found = df_ncm[df_ncm['Artículo'] == str(mat_code)]
    if len(found) == 0:
        return None, f"NCM no encontrado para material {mat_code}"
    return found.iloc[0]['NCM'], None

def verificar_origen_proximas(origen_anmat, mat_code, df_prox):
    if df_prox is None:
        return None, f"Material {mat_code}: no se pudo leer Próximas Importaciones"
    prox_row = df_prox[df_prox['Material'] == str(mat_code)]
    if len(prox_row) == 0:
        return None, f"Material {mat_code} no encontrado en Próximas Importaciones"
    origen_prox = str(prox_row.iloc[0]['Origen'])
    if not origen_prox or origen_prox == 'nan':
        return None, f"Material {mat_code}: origen vacío en Próximas Importaciones"
    pais_anmat = normalizar_pais(origen_anmat)
    pais_prox = normalizar_pais(origen_prox)
    if pais_anmat != pais_prox:
        return None, f"Origen no coincide para {mat_code}: ANMAT='{origen_anmat}' vs ProxImp='{origen_prox}'"
    return origen_prox, None

SEPARADORES_REGISTRO = [' - ', ' + ', ' / ', ' | ', '\n', '; ', '+', ',']

def separar_registros(registro_str):
    if not registro_str or registro_str == 'nan':
        return [registro_str]
    s = str(registro_str).strip()
    for sep in SEPARADORES_REGISTRO:
        if sep in s:
            partes = [p.strip() for p in s.split(sep) if p.strip()]
            if len(partes) > 1:
                return partes
    return [s]

def buscar_por_registro(nro_registro, df_anmat):
    nro = str(nro_registro).strip()
    found = df_anmat[df_anmat['Registros ANMAT'].apply(_codigo_str) == nro]
    if len(found) == 0:
        return None, "NOT_FOUND"
    if len(found) == 1:
        return [found.iloc[0]], None
    return [found.iloc[i] for i in range(len(found))], "MULTIPLE"

def parsear_fecha_vencimiento(expire_str):
    try:
        if not expire_str or expire_str == 'nan':
            return None
        partes = str(expire_str).strip().split('/')
        if len(partes) == 2:
            mes, anio = int(partes[0]), int(partes[1])
            return datetime(anio, mes, 1)
        return None
    except:
        return None

def verificar_vencimiento(expire_str):
    fecha = parsear_fecha_vencimiento(expire_str)
    if fecha is None:
        return 'ok', None
    hoy = datetime.now()
    limite_90 = hoy + timedelta(days=90)
    if fecha < hoy:
        return 'vencido', f"⚠️ VENCIDO: {expire_str}"
    if fecha <= limite_90:
        return 'proximo', f"⚠️ Vence próximo en 90 días: {expire_str}"
    return 'ok', None

def buscar_equivalente_en_bases(cod_equiv, df_anmat, df_avon, df_prox, df_fab, df_ncm, descripcion_pl=''):
    anmat_row = buscar_anmat(cod_equiv, df_anmat)
    if anmat_row is not None:
        nombre = str(anmat_row['NOMBRE']) if pd.notna(anmat_row['NOMBRE']) else ''
        variedad = str(anmat_row['Variedad']) if pd.notna(anmat_row['Variedad']) else ''
        contenido = str(anmat_row['CONTENIDO NETO']) if pd.notna(anmat_row['CONTENIDO NETO']) else ''
        registro = str(anmat_row['Registros ANMAT']) if pd.notna(anmat_row['Registros ANMAT']) else ''
        origen = str(anmat_row['ORIGEN']) if pd.notna(anmat_row['ORIGEN']) else ''
        if 'REFIL' in descripcion_pl.upper():
            nombre = nombre + ' (REPUESTO)'
        origen_norm = normalizar_pais(origen).capitalize() if origen != 'nan' else ''
        fab, _, _fab_row = buscar_fabricante(origen, cod_equiv, df_fab)
        ncm, _ = buscar_ncm(cod_equiv, df_ncm)
        return {'Marca y Nombre del producto': nombre if nombre != 'nan' else '', 'Variedades': variedad if variedad != 'nan' else '', 'Presentación': contenido if contenido != 'nan' else '', 'N° de inscripcion': registro if registro != 'nan' else '', 'Origen': origen_norm, 'Fabricante': fab or '', 'Posición Arancelaria': ncm or ''}, 'anmat', None
    avon_row = buscar_avon(cod_equiv, df_avon)
    if avon_row is not None:
        nombre_avon = str(avon_row.get('NOMBRE DE REGISTRO DE PRODUCTO', ''))
        contenido_avon = str(avon_row.get('CONTENIDO LEGAL', ''))
        registro_avon = str(avon_row.get('Reg. SP   (Trámite#)\nARGENTINA NATURA', ''))
        if 'REFIL' in descripcion_pl.upper():
            nombre_avon = nombre_avon + ' (REPUESTO)'
        ncm, _ = buscar_ncm(cod_equiv, df_ncm)
        return {'Marca y Nombre del producto': nombre_avon if nombre_avon != 'nan' else '', 'Variedades': '', 'Presentación': contenido_avon if contenido_avon != 'nan' else '', 'N° de inscripcion': registro_avon if registro_avon != 'nan' else '', 'Origen': '', 'Fabricante': '', 'Posición Arancelaria': ncm or ''}, 'avon', None
    return None, None, f"Código {cod_equiv} no encontrado en ANMAT ni Avon."

def procesar_pl(pl, df_anmat, df_avon, df_prox, df_fab, df_ncm):
    filas = []; alertas_excluir = []; alertas_avon = []; alertas_generales = []
    for _, pl_row in pl.iterrows():
        mat_code = str(pl_row[1]).strip()
        if not re.match(r'^\d{5,}$|^\d+-\d{4,}$', mat_code):
            continue
        cantidad_raw = pl_row[2]
        if pd.isna(cantidad_raw):
            cantidad = ''
        else:
            cantidad_str = str(cantidad_raw).strip()
            m_cant = re.match(r'^([\d,\.]+)', cantidad_str.replace(' ', ''))
            if m_cant:
                try: cantidad = int(float(m_cant.group(1).replace(',', '')))
                except: cantidad = cantidad_str
            else:
                cantidad = cantidad_str
        descripcion_pl = str(pl_row[3]).strip() if pd.notna(pl_row[3]) else ''
        lot_product = str(pl_row[5]).strip() if pd.notna(pl_row[5]) else ''
        expire_date = str(pl_row[6]).strip() if pd.notna(pl_row[6]) else ''
        fila = {'MATERIAL': mat_code, 'descripcion_factura': descripcion_pl, 'Marca y Nombre del producto': '', 'Variedades': '', 'Presentación': '', 'Cantidad': cantidad, 'N° de inscripcion': '', 'Lote': lot_product, 'Fecha de vencimiento': expire_date, 'Origen': '', 'Fabricante': '', 'Posición Arancelaria': '', '_alertas': [], '_skip': False, '_avon': False, '_necesita_completar': False, '_vencimiento': None, '_multi_registro': False, '_expanded': False}
        estado_venc, msg_venc = verificar_vencimiento(expire_date)
        if msg_venc:
            fila['_vencimiento'] = estado_venc; fila['_alertas'].append(msg_venc); alertas_generales.append(f"{mat_code} — {msg_venc}")
        anmat_row = buscar_anmat(mat_code, df_anmat)
        if anmat_row is not None:
            nombre = str(anmat_row['NOMBRE']) if pd.notna(anmat_row['NOMBRE']) else ''
            variedad = str(anmat_row['Variedad']) if pd.notna(anmat_row['Variedad']) else ''
            contenido = str(anmat_row['CONTENIDO NETO']) if pd.notna(anmat_row['CONTENIDO NETO']) else ''
            registro = str(anmat_row['Registros ANMAT']) if pd.notna(anmat_row['Registros ANMAT']) else ''
            origen = str(anmat_row['ORIGEN']) if pd.notna(anmat_row['ORIGEN']) else ''
            if 'REFIL' in descripcion_pl.upper():
                nombre = nombre + ' (REPUESTO)'
            fila['Origen'] = normalizar_pais(origen).capitalize() if origen != 'nan' else ''
            _, alerta_origen = verificar_origen_proximas(origen, mat_code, df_prox)
            if alerta_origen:
                fila['_alertas'].append(alerta_origen); alertas_generales.append(alerta_origen)
            fab, alerta_fab, _fab_row = buscar_fabricante(origen, mat_code, df_fab)
            if alerta_fab:
                fila['_alertas'].append(alerta_fab); alertas_generales.append(alerta_fab)
            else:
                fila['Fabricante'] = fab
            registros = separar_registros(registro)
            fila['Marca y Nombre del producto'] = nombre
            fila['Variedades'] = variedad if variedad != 'nan' else ''
            fila['Presentación'] = contenido if contenido != 'nan' else ''
            fila['N° de inscripcion'] = registro if registro != 'nan' else ''
            if len(registros) > 1:
                fila['_multi_registro'] = True
                idx_fila_principal = len(filas)
                filas.append(fila)
                for nro in registros:
                    anmat_rows, status = buscar_por_registro(nro, df_anmat)
                    if status == "NOT_FOUND":
                        fila_exp = {'MATERIAL': '', 'descripcion_factura': '', 'Marca y Nombre del producto': '', 'Variedades': '', 'Presentación': '', 'Cantidad': '', 'N° de inscripcion': nro, 'Lote': '', 'Fecha de vencimiento': '', 'Origen': '', 'Fabricante': '', 'Posición Arancelaria': '', '_alertas': ["No encontrado: " + nro], '_skip': False, '_avon': False, '_necesita_completar': False, '_expanded': True, '_multi_opciones': False, '_nro_registro': nro}
                        alertas_generales.append("No encontrado: " + nro); filas.append(fila_exp)
                    else:
                        es_multiple = status == "MULTIPLE"
                        for multi_i, anmat_nro in enumerate(anmat_rows):
                            n = str(anmat_nro['NOMBRE']) if pd.notna(anmat_nro['NOMBRE']) else ''
                            v = str(anmat_nro['Variedad']) if pd.notna(anmat_nro['Variedad']) else ''
                            c = str(anmat_nro['CONTENIDO NETO']) if pd.notna(anmat_nro['CONTENIDO NETO']) else ''
                            if 'REFIL' in descripcion_pl.upper(): n = n + ' (REPUESTO)'
                            fila_exp = {'MATERIAL': '', 'descripcion_factura': '', 'Marca y Nombre del producto': n, 'Variedades': v if v != 'nan' else '', 'Presentación': c if c != 'nan' else '', 'Cantidad': '', 'N° de inscripcion': nro, 'Lote': '', 'Fecha de vencimiento': '', 'Origen': '', 'Fabricante': '', 'Posición Arancelaria': '', '_alertas': [], '_skip': es_multiple, '_avon': False, '_necesita_completar': False, '_expanded': True, '_multi_opciones': es_multiple, '_nro_registro': nro, '_multi_idx': multi_i}
                            filas.append(fila_exp)
                ncm, alerta_ncm = buscar_ncm(mat_code, df_ncm)
                if alerta_ncm: alertas_generales.append(alerta_ncm)
                else: filas[idx_fila_principal]['Posición Arancelaria'] = ncm
                continue
        else:
            avon_row = buscar_avon(mat_code, df_avon)
            if avon_row is not None:
                fila['_avon'] = True
                def _get_avon(row, variantes, default=''):
                    for v in variantes:
                        val = row.get(v)
                        if val is not None: return str(val).strip()
                    return default
                nombre_avon = _get_avon(avon_row, ['NOMBRE DE REGISTRO DE PRODUCTO', 'NOMBRE REGISTRO', 'NOMBRE'])
                contenido_avon = _get_avon(avon_row, ['CONTENIDO LEGAL', 'CONTENIDO'])
                registro_avon = _get_avon(avon_row, ['Reg. SP   (Trámite#)\nARGENTINA NATURA', 'Reg. SP   (Trámite#)\nNATURA ARG', 'Reg. SP (Trámite#)\nARGENTINA NATURA', 'Reg. SP   (Tramite#)\nARGENTINA NATURA', 'Reg. SP   (Trámite#)\nNATURA ARGENTINA'])
                if 'REFIL' in descripcion_pl.upper(): nombre_avon = nombre_avon + ' (REPUESTO)'
                fila['Marca y Nombre del producto'] = nombre_avon if nombre_avon != 'nan' else ''
                fila['Presentación'] = contenido_avon if contenido_avon != 'nan' else ''
                fila['N° de inscripcion'] = registro_avon if registro_avon != 'nan' else ''
                fila['Variedades'] = ''; fila['Origen'] = ''; fila['_necesita_completar'] = True
                fab, alerta_fab, _fab_row = buscar_fabricante('', mat_code, df_fab)
                fila['Fabricante'] = fab if not alerta_fab else ''
                avon_idx_actual = len(alertas_avon)
                alertas_avon.append({'material': mat_code, 'descripcion': descripcion_pl, 'fila_idx': len(filas), 'avon_idx': avon_idx_actual})
                fila['_avon_idx'] = avon_idx_actual
            else:
                fila['_skip'] = True; fila['_no_encontrado'] = True
                alertas_excluir.append({'material': mat_code, 'descripcion': descripcion_pl, 'fila_idx': len(filas)})
        ncm, alerta_ncm = buscar_ncm(mat_code, df_ncm)
        if alerta_ncm: fila['_alertas'].append(alerta_ncm); alertas_generales.append(alerta_ncm); fila['Posición Arancelaria'] = ''
        else: fila['Posición Arancelaria'] = ncm
        filas.append(fila)
    return filas, alertas_excluir, alertas_avon, alertas_generales

def separar_anexos(filas):
    principal, difusor, muestras, alertas_sep = [], [], [], []
    palabras_muestra = ['amostra', 'muestra', 'sample', 'muestras', 'amostras']
    for fila in filas:
        if fila['_skip']: continue
        desc = fila['descripcion_factura'].upper()
        desc_lower = fila['descripcion_factura'].lower()
        es_difusor = 'DIFUSOR' in desc
        es_3x1 = bool(re.search(r'3\s*[Xx]\s*1(?![0-9])', desc))
        es_muestra = any(p in desc_lower for p in palabras_muestra)
        if es_difusor and es_3x1: alertas_sep.append(f"Material {fila['MATERIAL']} tiene DIFUSOR y 3X1 — verificar."); principal.append(fila)
        elif es_difusor: difusor.append(fila)
        elif es_3x1 and es_muestra: muestras.append(fila)
        else: principal.append(fila)
    return principal, difusor, muestras, alertas_sep

FABRICANTE_MUESTRAS = 'INDUSTRIA E COMERCIO DE COSMÉTICOS NATURA LTDA'
ORIGEN_MUESTRAS = 'Brasil'

def parsear_msg(file_bytes):
    try:
        import extract_msg
        with tempfile.NamedTemporaryFile(suffix='.msg', delete=False) as f:
            f.write(file_bytes); tmp = f.name
        msg = extract_msg.Message(tmp); body = msg.body or ''
    except Exception as e:
        return None, f"No se pudo leer el archivo .msg: {e}"
    primer_bloque = re.split(r'_{3,}', body)[0]
    lineas = [l.strip() for l in primer_bloque.replace('\r\n', '\n').replace('\r', '\n').split('\n')]
    inicio = None
    for i, l in enumerate(lineas):
        if 'código del artículo' in l.lower() or 'codigo del articulo' in l.lower():
            inicio = i + 1; break
    if inicio is not None:
        cols_ignorar = {'ncm', 'anmat', 'código del artículo', 'codigo del articulo', ''}
        tokens = [l for l in lineas[inicio:] if l.lower() not in cols_ignorar]
        items = []; i = 0
        while i + 2 < len(tokens):
            codigo = tokens[i].strip(); ncm = tokens[i+1].strip(); anmat_val = tokens[i+2].strip().lower()
            if anmat_val in ('si', 'sí', 'no') and re.match(r'^\d{4,}', ncm):
                items.append({'codigo': codigo, 'ncm': ncm, 'anmat': anmat_val in ('si', 'sí')}); i += 3
            else: i += 1
        if items: return items, None
    return None, "No se encontró la tabla Código / NCM / ANMAT en el mail."

def cargar_clasificacion_excel(file_bytes):
    try:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            f.write(file_bytes); tmp = f.name
        df = pd.read_excel(tmp, header=0)
    except Exception as e:
        return None, f"No se pudo leer el archivo Excel: {e}"
    col_map = {c.strip().lower(): c for c in df.columns}
    def _buscar_col(variantes):
        for v in variantes:
            if v in col_map: return col_map[v]
        return None
    col_material = _buscar_col(['material', 'código', 'codigo'])
    col_ncm = _buscar_col(['ncm'])
    col_observ = _buscar_col(['observ.', 'observ', 'observaciones', 'anmat'])
    if col_material is None or col_ncm is None:
        return None, "El Excel debe tener al menos las columnas Material y NCM."
    items = []
    for _, row in df.iterrows():
        codigo = str(row[col_material]).strip() if pd.notna(row[col_material]) else ''
        if not re.match(r'^\d{5,}|^\d+-\d+', codigo): continue
        ncm = str(row[col_ncm]).strip() if pd.notna(row[col_ncm]) else ''
        observ = str(row[col_observ]).strip() if col_observ and pd.notna(row[col_observ]) else ''
        es_anmat = 'anmat' in observ.lower()
        items.append({'codigo': codigo, 'ncm': ncm, 'anmat': es_anmat})
    if not items: return None, "No se encontraron filas válidas en el Excel de clasificación."
    return items, None

def cargar_pl_muestras(file_bytes):
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        f.write(file_bytes); tmp = f.name
    xl = pd.ExcelFile(tmp); invoice = None; items = []
    for sh in xl.sheet_names:
        df_raw = pd.read_excel(tmp, sheet_name=sh, header=None)
        if not invoice:
            for i, row in df_raw.iterrows():
                vals = list(row.values)
                for j, val in enumerate(vals):
                    val_str = str(val).replace('\n', ' ')
                    if 'Nº INVOICE:' in val_str or 'N° INVOICE:' in val_str:
                        parte = val_str.split(':', 1)[1].strip()
                        if parte and parte != 'nan': invoice = parte; break
                        for k in range(j + 1, len(vals)):
                            v = str(vals[k]).strip()
                            if v and v != 'nan': invoice = v; break
                        break
                if invoice: break
        header_row_idx = None
        for i, row in df_raw.iterrows():
            if any('MATERIAL CODE' in str(v).upper() for v in row.values if v):
                header_row_idx = i; break
        if header_row_idx is None: continue
        headers = [str(v).strip().upper() if v else '' for v in df_raw.iloc[header_row_idx].values]
        col_material = col_qty_kg = col_qty_g = col_qty_un = col_desc = col_lot = col_expire = None
        for idx, h in enumerate(headers):
            if 'MATERIAL CODE' in h: col_material = idx
            elif ('QUANTITY' in h or 'CANTIDAD' in h) and ('KG' in h or 'KILO' in h):
                if col_qty_kg is None: col_qty_kg = idx
            elif ('QUANTITY' in h or 'CANTIDAD' in h) and ('GRAM' in h) and 'KG' not in h:
                if col_qty_g is None: col_qty_g = idx
            elif ('QUANTITY' in h or 'CANTIDAD' in h) and ('UNIT' in h or 'UNID' in h):
                if col_qty_un is None: col_qty_un = idx
            elif ('QUANTITY' in h or 'CANTIDAD' in h) and col_qty_kg is None and col_qty_g is None and col_qty_un is None:
                col_qty_kg = idx
            elif 'DESCRIPTION' in h or 'DESCRIPCI' in h:
                if col_desc is None: col_desc = idx
            elif 'LOT' in h or 'LOTE' in h:
                if col_lot is None: col_lot = idx
            elif 'EXPIRE' in h or 'VENC' in h:
                if col_expire is None: col_expire = idx
        if col_material is None: continue
        data_start = header_row_idx + 2
        for i in range(data_start, len(df_raw)):
            row = df_raw.iloc[i]
            mat = str(row.iloc[col_material]).strip() if col_material is not None else ''
            if not mat or mat == 'nan' or not re.search(r'\d', mat): continue
            if any(kw in mat.upper() for kw in ('VOLUME', 'OBSERVAC', 'TOTAL', 'EXEMPLO', 'INFORM', 'PALLET', 'FUMIGATE', 'BOX', 'ISPM')): continue
            if not re.search(r'\d{4,}', mat) and not re.match(r'^\d+-\d+', mat): continue
            desc = str(row.iloc[col_desc]).strip() if col_desc is not None and pd.notna(row.iloc[col_desc]) else ''
            lot = str(row.iloc[col_lot]).strip() if col_lot is not None and pd.notna(row.iloc[col_lot]) else ''
            expire_raw = row.iloc[col_expire] if col_expire is not None else None
            if isinstance(expire_raw, datetime): expire_str = expire_raw.strftime('%m/%Y')
            elif expire_raw is not None and str(expire_raw) != 'nan': expire_str = str(expire_raw).strip()
            else: expire_str = ''
            cantidades = {}
            for tipo, col in [('kg', col_qty_kg), ('g', col_qty_g), ('un', col_qty_un)]:
                if col is not None:
                    v = row.iloc[col]
                    if pd.notna(v) and str(v).strip() not in ('nan', ''):
                        try: cantidades[tipo] = float(v)
                        except: pass
            items.append({'material': mat, 'descripcion': desc, 'lot': lot, 'expire': expire_str, 'cantidades': cantidades})
    return items, invoice

def procesar_muestras(items_pl, items_mail):
    sufijo_map = {'kg': 'kg', 'g': 'g', 'un': 'un'}
    col_map = {'kg': 'Cantidad en KG', 'g': 'Cantidad en gramos', 'un': 'Cantidad en unidades'}
    nombre_pres = {'kg': 'kilos', 'g': 'gramos', 'un': 'unidades'}
    mail_dict = {it['codigo'].strip(): it for it in items_mail}
    codigos_si = {cod for cod, it in mail_dict.items() if it['anmat']}
    codigos_pl = {it['material'] for it in items_pl}
    tipos_global = set()
    for item in items_pl:
        if item['material'] in codigos_si: tipos_global.update(item['cantidades'].keys())
    un_solo_tipo = len(tipos_global) == 1
    tipo_unico = list(tipos_global)[0] if un_solo_tipo else None
    col_cantidad_header = col_map.get(tipo_unico, 'Cantidad') if un_solo_tipo else 'Cantidad'
    filas = []
    for item in items_pl:
        mat = item['material']
        if mat not in codigos_si: continue
        mail_item = mail_dict[mat]; cantidades = item['cantidades']
        tipos_item = [t for t in ('kg', 'g', 'un') if t in cantidades]
        if len(tipos_item) == 0: presentacion = ''; cantidad_val = ''
        elif len(tipos_item) == 1:
            t = tipos_item[0]; presentacion = nombre_pres[t]; v = cantidades[t]
            if un_solo_tipo: cantidad_val = int(v) if v == int(v) else v
            else: cantidad_val = f"{int(v) if v == int(v) else v} {sufijo_map[t]}"
        else:
            presentacion = ' / '.join(nombre_pres[t] for t in tipos_item)
            partes = [f"{int(cantidades[t]) if cantidades[t] == int(cantidades[t]) else cantidades[t]} {sufijo_map[t]}" for t in tipos_item]
            cantidad_val = ' / '.join(partes)
        filas.append({'MATERIAL': mat, 'descripcion_factura': item['descripcion'], 'Marca y Nombre del producto': item['descripcion'], 'Variedades': 'N/C', 'Presentación': presentacion, 'Cantidad': cantidad_val, 'N° de inscripcion': 'N/C', 'Lote': item['lot'], 'Fecha de vencimiento': item['expire'], 'Origen': ORIGEN_MUESTRAS, 'Fabricante': FABRICANTE_MUESTRAS, 'Posición Arancelaria': mail_item['ncm'], '_alertas': [], '_skip': False, '_avon': False, '_necesita_completar': False, '_vencimiento': None, '_multi_registro': False, '_expanded': False})
    no_en_pl = [cod for cod in codigos_si if cod not in codigos_pl]
    return filas, no_en_pl, col_cantidad_header

COLUMNAS_SALIDA = ['MATERIAL', 'descripcion_factura', 'Marca y Nombre del producto', 'Variedades', 'Presentación', 'Cantidad', 'N° de inscripcion', 'Lote', 'Fecha de vencimiento', 'Origen', 'Fabricante', 'Posición Arancelaria']
COLUMNAS_SIN_PRIMERAS = COLUMNAS_SALIDA[2:]
ANCHOS = {'MATERIAL': 14, 'descripcion_factura': 38, 'Marca y Nombre del producto': 48, 'Variedades': 14, 'Presentación': 14, 'Cantidad': 14, 'N° de inscripcion': 24, 'Lote': 18, 'Fecha de vencimiento': 20, 'Origen': 14, 'Fabricante': 48, 'Posición Arancelaria': 22}
LEYENDA_ROTULADO = "será sobrerotulado en depósito con los datos legales exigidos por la normativa Argentina previo a su comercialización."

def escribir_excel_bytes(filas, incluir_primeras_cols=True, col_cantidad_header='Cantidad', materiales_rotulado=None):
    wb = Workbook(); ws = wb.active; ws.title = 'Anexo de Productos'
    columnas = COLUMNAS_SALIDA if incluir_primeras_cols else COLUMNAS_SIN_PRIMERAS
    ws.merge_cells(f'A1:{get_column_letter(len(columnas))}1')
    titulo = ws['A1']; titulo.value = 'ANEXO DE PRODUCTOS'
    titulo.font = Font(name='Arial', bold=True, size=11); titulo.alignment = Alignment(horizontal='center', vertical='center'); titulo.fill = PatternFill('solid', start_color='D9D9D9')
    header_fill = PatternFill('solid', start_color='70AD47')
    for col_idx, col_name in enumerate(columnas, 1):
        display_name = col_cantidad_header if col_name == 'Cantidad' else col_name
        cell = ws.cell(row=2, column=col_idx, value=display_name)
        cell.font = Font(name='Arial', bold=True, size=11, color='FFFFFF'); cell.fill = header_fill; cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    alerta_fill = PatternFill('solid', start_color='FFEB9C')
    for row_idx, fila in enumerate(filas, 3):
        tiene_alerta = len(fila.get('_alertas', [])) > 0 or fila.get('_necesita_completar', False)
        for col_idx, col_name in enumerate(columnas, 1):
            val = fila.get(col_name, '')
            if col_name == 'MATERIAL' and val != '':
                try: val = int(float(str(val)))
                except: pass
            if col_name == 'Cantidad' and val != '':
                if not isinstance(val, str):
                    try:
                        v = float(val); val = int(v) if v == int(v) else v
                    except: pass
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name='Calibri', size=11); cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if tiene_alerta: cell.fill = alerta_fill
    if materiales_rotulado:
        ultima_fila_datos = len(filas) + 2; fila_leyenda = ultima_fila_datos + 2
        col_leyenda_idx = 4 if incluir_primeras_cols else 2
        mats_str = ', '.join(str(m) for m in materiales_rotulado)
        texto_leyenda = f"Material {mats_str}: {LEYENDA_ROTULADO}"
        col_inicio_letter = get_column_letter(col_leyenda_idx); col_fin_letter = get_column_letter(len(columnas))
        ws.merge_cells(f'{col_inicio_letter}{fila_leyenda}:{col_fin_letter}{fila_leyenda}')
        cell_leyenda = ws.cell(row=fila_leyenda, column=col_leyenda_idx, value=texto_leyenda)
        cell_leyenda.font = Font(name='Calibri', size=10, bold=True); cell_leyenda.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True); ws.row_dimensions[fila_leyenda].height = 30
    for col_idx, col_name in enumerate(columnas, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = ANCHOS.get(col_name, 15)
    ws.row_dimensions[1].height = 22; ws.row_dimensions[2].height = 40; ws.freeze_panes = 'A3'
    if not incluir_primeras_cols:
        ws.page_setup.orientation = 'landscape'; ws.page_setup.fitToPage = True; ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
        ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5)
    buf = BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()

def excel_a_pdf_bytes(excel_bytes, nombre_base):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        wb = load_workbook(BytesIO(excel_bytes)); ws = wb.active
        style_normal = ParagraphStyle('normal', fontSize=6.5, leading=8, alignment=TA_CENTER)
        style_header = ParagraphStyle('header', fontSize=7, leading=9, alignment=TA_CENTER, textColor=colors.white, fontName='Helvetica-Bold')
        style_title = ParagraphStyle('title', fontSize=8, leading=10, alignment=TA_CENTER, fontName='Helvetica-Bold')
        style_leyenda = ParagraphStyle('leyenda', fontSize=7, leading=9, alignment=TA_LEFT, fontName='Helvetica-Bold')
        ancho_total = landscape(A4)[0] - 1.4*cm
        pesos = [4.0, 1.4, 1.4, 1.2, 2.2, 1.4, 1.8, 1.2, 3.8, 2.0]
        total_pesos = sum(pesos); col_widths = [ancho_total * p / total_pesos for p in pesos]
        def safe_para(val, style):
            try:
                txt = str(val) if val is not None else ''
                txt = txt.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                return Paragraph(txt, style)
            except: return Paragraph('', style)
        data = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0: data.append([safe_para(row[0], style_title)] + ['' for _ in range(len(pesos)-1)])
            elif row_idx == 1:
                cells = list(row)
                while len(cells) < len(pesos): cells.append('')
                cells = cells[:len(pesos)]; data.append([safe_para(c, style_header) for c in cells])
            else:
                cells = list(row); fila_str = ' '.join(str(c) for c in cells if c)
                if LEYENDA_ROTULADO[:20] in fila_str:
                    texto_leyenda = next((str(c) for c in cells if c and LEYENDA_ROTULADO[:20] in str(c)), '')
                    data.append([safe_para(texto_leyenda, style_leyenda)] + ['' for _ in range(len(pesos)-1)])
                else:
                    while len(cells) < len(pesos): cells.append('')
                    cells = cells[:len(pesos)]; data.append([safe_para(c, style_normal) for c in cells])
        if not data: return None
        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=0.7*cm, rightMargin=0.7*cm, topMargin=0.8*cm, bottomMargin=0.8*cm)
        table = Table(data, colWidths=col_widths, repeatRows=2)
        n_filas = len(data)
        style_cmds = [('BACKGROUND', (0,0), (-1,0), colors.HexColor('#D9D9D9')), ('SPAN', (0,0), (-1,0)), ('BACKGROUND', (0,1), (-1,1), colors.HexColor('#70AD47')), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#CCCCCC')), ('ROWBACKGROUNDS', (0,2), (-1,-1), [colors.white, colors.HexColor('#F7F7F7')]), ('LEFTPADDING', (0,0), (-1,-1), 3), ('RIGHTPADDING', (0,0), (-1,-1), 3), ('TOPPADDING', (0,0), (-1,-1), 2), ('BOTTOMPADDING', (0,0), (-1,-1), 2)]
        if n_filas > 2:
            last_row = data[-1]; fila_str_last = str(last_row[0]) if last_row else ''
            if LEYENDA_ROTULADO[:15] in fila_str_last or 'sobrerotulado' in fila_str_last:
                style_cmds.append(('SPAN', (0, n_filas-1), (-1, n_filas-1))); style_cmds.append(('ALIGN', (0, n_filas-1), (-1, n_filas-1), 'LEFT'))
        table.setStyle(TableStyle(style_cmds)); doc.build([table]); buf.seek(0); return buf.getvalue()
    except Exception as e:
        print(f'PDF error: {e}'); return None

def generar_zip(grupos, invoice, col_cantidad_header='Cantidad', materiales_rotulado=None):
    buf = BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for nombre, filas in grupos:
            if not filas: continue
            nombre_base = f'ANEXO_{nombre}_{invoice}'
            xls_completo = escribir_excel_bytes(filas, incluir_primeras_cols=True, col_cantidad_header=col_cantidad_header, materiales_rotulado=materiales_rotulado)
            zf.writestr(f'{nombre_base}.xlsx', xls_completo)
            xls_sin = escribir_excel_bytes(filas, incluir_primeras_cols=False, col_cantidad_header=col_cantidad_header, materiales_rotulado=materiales_rotulado)
            zf.writestr(f'{nombre_base}_SIN_MAT.xlsx', xls_sin)
            pdf = excel_a_pdf_bytes(xls_sin, f'{nombre_base}_SIN_MAT')
            if pdf: zf.writestr(f'{nombre_base}_SIN_MAT.pdf', pdf)
    buf.seek(0); return buf.getvalue()

# SESSION STATE
defaults = {'filas_procesadas': None, 'alertas_excluir': [], 'alertas_avon': [], 'alertas_generales': [], 'invoice': None, 'excluidos': set(), 'alerta_origen_proveedor': None, 'datos_avon_completados': {}, 'df_avon_editable': None, '_avon_init_invoice': None, 'filas_muestras': None, 'invoice_muestras': None, 'alertas_muestras': [], 'col_cantidad_muestras': 'Cantidad', 'fiabila_coas': [], 'equivalentes': {}, 'rotulado_activo': False, 'materiales_rotulado': [], '_df_anmat_cache': None, '_df_avon_cache': None, '_df_fab_cache': None, '_df_ncm_cache': None}
for k, v in defaults.items():
    if k not in st.session_state: st.session_state[k] = v

st.markdown('<div class="card"><h3><span class="step-badge">0</span>Tipo de operación</h3>', unsafe_allow_html=True)
modo = st.radio("¿Qué tipo de operación es?", options=["Operación normal", "Muestras Natura", "Fiabila"], horizontal=True, key='p6_modo_radio')
st.markdown('</div>', unsafe_allow_html=True)

if modo == "Muestras Natura":
    st.markdown('<div class="modo-muestras">🧪 Modo Muestras Natura</div>', unsafe_allow_html=True)
    st.markdown('<div class="card"><h3><span class="step-badge">1</span>Archivos de la operación</h3>', unsafe_allow_html=True)
    st.markdown("**📌 Número de referencia de la operación**")
    nro_ref_m = st.text_input("Número de referencia muestras", placeholder="ej: MN014-26", label_visibility="collapsed", key='p6_nro_ref_muestras')
    col1, col2 = st.columns(2)
    with col1: f_pl_m = st.file_uploader("📦 Packing List / Invoice (.xlsx)", type=['xlsx'], key='p6_pl_muestras')
    with col2: f_msg = st.file_uploader("📧 Clasificación ANMAT (.msg ó .xlsx)", type=['msg', 'xlsx'], key='p6_msg_muestras')
    st.markdown('</div>', unsafe_allow_html=True)
    if f_pl_m and f_msg:
        st.markdown('<div class="card"><h3><span class="step-badge">2</span>Procesar</h3>', unsafe_allow_html=True)
        if st.button("⚙️ Analizar y procesar muestras", key='p6_btn_procesar_muestras'):
            with st.spinner('Procesando...'):
                try:
                    if f_msg.name.lower().endswith('.xlsx'): items_mail, err_mail = cargar_clasificacion_excel(f_msg.read())
                    else: items_mail, err_mail = parsear_msg(f_msg.read())
                    if err_mail: st.error(f"Error al leer la clasificación: {err_mail}"); st.stop()
                    items_pl, invoice_m = cargar_pl_muestras(f_pl_m.read())
                    if not items_pl: st.error("No se encontraron ítems en el Packing List."); st.stop()
                    filas_m, no_en_pl, col_cant_hdr = procesar_muestras(items_pl, items_mail)
                    alertas_m = [f"⚠️ Código {cod} tiene ANMAT=Sí en el mail pero no se encontró en el Packing List." for cod in no_en_pl]
                    st.session_state.filas_muestras = filas_m; st.session_state.invoice_muestras = invoice_m; st.session_state.alertas_muestras = alertas_m; st.session_state.col_cantidad_muestras = col_cant_hdr
                except Exception as e:
                    import traceback; st.error(f"Error al procesar: {e}"); st.text(traceback.format_exc())
        st.markdown('</div>', unsafe_allow_html=True)
    if st.session_state.filas_muestras is not None:
        filas_m = st.session_state.filas_muestras; invoice_m = st.session_state.invoice_muestras
        col1, col2 = st.columns(2)
        with col1: st.markdown(f'<div class="stat-card"><div class="number">{len(filas_m)}</div><div class="label">Ítems con ANMAT = Sí</div></div>', unsafe_allow_html=True)
        with col2: st.markdown(f'<div class="stat-card"><div class="number" style="color:#00c896">{len(filas_m)}</div><div class="label">Líneas en el Anexo</div></div>', unsafe_allow_html=True)
        st.markdown('<br>', unsafe_allow_html=True)
        if st.session_state.alertas_muestras:
            st.markdown('<div class="card"><h3>⚠️ Alertas</h3>', unsafe_allow_html=True)
            for a in st.session_state.alertas_muestras: st.markdown(f'<div class="alert-box">{a}</div>', unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)
        st.markdown('<div class="card"><h3><span class="step-badge">4</span>Generar Anexo de Muestras</h3>', unsafe_allow_html=True)
        if st.button("📄 Generar Anexo de Muestras", key='p6_btn_generar_muestras'):
            with st.spinner('Generando archivos...'):
                ref = nro_ref_m.strip() if nro_ref_m.strip() else (invoice_m or 'MUESTRAS')
                col_hdr = st.session_state.get('col_cantidad_muestras', 'Cantidad')
                zip_bytes = generar_zip([('MUESTRAS', filas_m)], ref, col_cantidad_header=col_hdr)
                st.markdown('<div class="success-box">✅ Anexo de Muestras generado correctamente</div>', unsafe_allow_html=True)
                st.download_button(label="⬇️ Descargar Anexo de Muestras (ZIP)", data=zip_bytes, file_name=f"ANEXO_MUESTRAS_{ref}.zip", mime="application/zip", key='p6_dl_zip_muestras')
        st.markdown('</div>', unsafe_allow_html=True)

elif modo in ("Operación normal", "Fiabila"):
    st.markdown('<div class="card"><h3><span class="step-badge">1</span>Archivos de la operación</h3>', unsafe_allow_html=True)
    st.markdown("**📌 Número de referencia de la operación**")
    nro_referencia = st.text_input("Número de referencia", placeholder="ej: 4550595912", label_visibility="collapsed")
    st.markdown("**🏷️ Rotulado**")
    rotulado_opcion = st.radio("¿Algún artículo tiene rotulado?", options=["No", "Sí"], horizontal=True, key='p6_rotulado_radio')
    st.session_state.rotulado_activo = (rotulado_opcion == "Sí")
    col1, col2 = st.columns(2)
    with col1:
        f_pl = st.file_uploader("📦 Packing List", type=['xlsx'], key='p6_pl')
        f_prox = st.file_uploader("📅 Próximas Importaciones", type=['xlsx', 'pdf'], key='p6_prox')
        f_anmat = st.file_uploader("🏥 Registro ANMAT Histórico", type=['xlsb','xlsx'], key='p6_anmat')
    with col2:
        f_avon = st.file_uploader("🌸 Registros Avon", type=['xlsx'], key='p6_avon')
        f_fab = st.file_uploader("🏭 Fabricantes", type=['xls','xlsx'], key='p6_fab')
        f_ncm = st.file_uploader("📊 Catálogo NCM", type=['xlsx'], key='p6_ncm')
    f_coas = []
    if modo == "Fiabila":
        f_coas = st.file_uploader("📄 COA(s) PDF", type=['pdf'], accept_multiple_files=True, key='p6_coas_fiabila')
    st.markdown('</div>', unsafe_allow_html=True)
    archivos_ok = all([f_pl, f_anmat, f_avon, f_fab, f_ncm]) and ((modo == "Operación normal" and f_prox) or (modo == "Fiabila" and len(f_coas) > 0))
    if archivos_ok:
        st.markdown('<div class="card"><h3><span class="step-badge">2</span>Procesar operación</h3>', unsafe_allow_html=True)
        if st.button("⚙️ Analizar y procesar", key='p6_btn_procesar'):
            with st.spinner('Procesando...'):
                try:
                    suffix_fab = '.xls' if f_fab.name.endswith('.xls') else '.xlsx'
                    df_avon = cargar_avon(f_avon.read()); df_fab = cargar_fabricantes(f_fab.read(), suffix=suffix_fab); df_ncm = cargar_ncm(f_ncm.read())
                    df_anmat = cargar_anmat(f_anmat.read())
                    df_prox, es_pdf_prox, origen_explicito_prox, origen_proveedor_prox = cargar_proximas(f_prox.read(), f_prox.name)
                    if es_pdf_prox and not origen_explicito_prox and origen_proveedor_prox: st.session_state.alerta_origen_proveedor = origen_proveedor_prox
                    else: st.session_state.alerta_origen_proveedor = None
                    pl, invoice = cargar_pl(f_pl.read())
                    st.session_state._df_anmat_cache = df_anmat; st.session_state._df_avon_cache = df_avon; st.session_state._df_fab_cache = df_fab; st.session_state._df_ncm_cache = df_ncm
                    filas, alertas_excluir, alertas_avon, alertas_generales = procesar_pl(pl, df_anmat, df_avon, df_prox, df_fab, df_ncm)
                    st.session_state.filas_procesadas = filas; st.session_state.alertas_excluir = alertas_excluir; st.session_state.alertas_avon = alertas_avon; st.session_state.alertas_generales = alertas_generales; st.session_state.invoice = invoice; st.session_state.excluidos = set(); st.session_state.equivalentes = {}; st.session_state.rotulado_activo = False; st.session_state.materiales_rotulado = []
                except Exception as e:
                    import traceback; st.error(f"Error al procesar: {e}"); st.text(traceback.format_exc())
        st.markdown('</div>', unsafe_allow_html=True)
    if st.session_state.filas_procesadas is not None:
        filas = st.session_state.filas_procesadas; invoice = st.session_state.invoice
        total = len(filas); skip = len(st.session_state.alertas_excluir); avon = len(st.session_state.alertas_avon); generales = len(st.session_state.alertas_generales)
        col1, col2, col3, col4 = st.columns(4)
        with col1: st.markdown(f'<div class="stat-card"><div class="number">{total}</div><div class="label">Ítems PL</div></div>', unsafe_allow_html=True)
        with col2: st.markdown(f'<div class="stat-card"><div class="number" style="color:#00c896">{total - skip}</div><div class="label">A procesar</div></div>', unsafe_allow_html=True)
        with col3: st.markdown(f'<div class="stat-card"><div class="number" style="color:#ffd166">{avon}</div><div class="label">Avon / completar</div></div>', unsafe_allow_html=True)
        with col4: st.markdown(f'<div class="stat-card"><div class="number" style="color:#ff6b6b">{skip}</div><div class="label">No encontrados</div></div>', unsafe_allow_html=True)
        st.markdown('<br>', unsafe_allow_html=True)
        st.markdown('<div class="card"><h3><span class="step-badge">3</span>Generar Anexo</h3>', unsafe_allow_html=True)
        if st.button("📄 Generar Anexo completo", key='p6_btn_generar'):
            with st.spinner('Generando archivos...'):
                filas_final = []
                for fila in filas:
                    mat = fila['MATERIAL']
                    if mat in st.session_state.excluidos: continue
                    if fila.get('_skip'):
                        equiv = st.session_state.equivalentes.get(mat, {})
                        if equiv.get('datos') is not None and mat not in st.session_state.excluidos:
                            f = fila.copy(); f['_skip'] = False; datos_eq = equiv['datos']
                            f['Marca y Nombre del producto'] = datos_eq.get('Marca y Nombre del producto', ''); f['Variedades'] = datos_eq.get('Variedades', ''); f['Presentación'] = datos_eq.get('Presentación', ''); f['N° de inscripcion'] = datos_eq.get('N° de inscripcion', ''); f['Origen'] = datos_eq.get('Origen', ''); f['Fabricante'] = datos_eq.get('Fabricante', ''); f['Posición Arancelaria'] = datos_eq.get('Posición Arancelaria', '')
                        else: continue
                    else: f = fila.copy()
                    filas_final.append(f)
                principal, difusor, muestras, _ = separar_anexos(filas_final)
                grupos = [('PRINCIPAL', principal), ('DIFUSOR', difusor), ('MUESTRAS', muestras)]
                ref = nro_referencia.strip() if nro_referencia.strip() else invoice
                buf_zip = BytesIO()
                with zipfile.ZipFile(buf_zip, 'w', zipfile.ZIP_DEFLATED) as zf:
                    for nombre_grupo, filas_grupo in grupos:
                        if not filas_grupo: continue
                        nombre_base = f'ANEXO_{nombre_grupo}_{ref}'
                        xls_completo = escribir_excel_bytes(filas_grupo, incluir_primeras_cols=True)
                        zf.writestr(f'{nombre_base}.xlsx', xls_completo)
                        xls_sin = escribir_excel_bytes(filas_grupo, incluir_primeras_cols=False)
                        zf.writestr(f'{nombre_base}_SIN_MAT.xlsx', xls_sin)
                        pdf = excel_a_pdf_bytes(xls_sin, f'{nombre_base}_SIN_MAT')
                        if pdf: zf.writestr(f'{nombre_base}_SIN_MAT.pdf', pdf)
                buf_zip.seek(0)
                st.markdown('<div class="success-box">✅ Anexo generado correctamente</div>', unsafe_allow_html=True)
                st.download_button(label="⬇️ Descargar todos los archivos (ZIP)", data=buf_zip.getvalue(), file_name=f"ANEXO_{ref}.zip", mime="application/zip", key='p6_dl_zip')
        st.markdown('</div>', unsafe_allow_html=True)
